# -*- coding: utf-8 -*-
"""
Created on Wed Nov  3 11:07:44 2021

@author: AARON.RAMIREZ
"""

import traceback
import numpy as np
import openpyxl as op
import pandas as pd
import string
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font, Alignment
from collections import Counter
from catalogos import validar_catalogo
from extractor import Extractor, Datos

frame = Datos('preguntas_relacionadas')
x_valida = Extractor()
final_tabla = Extractor()

redFill = PatternFill(start_color='FF0000',end_color='FF0000',fill_type='solid')

gris = PatternFill(fill_type='mediumGray',patternType='mediumGray')

azul = PatternFill(start_color='008B8B',end_color='008B8B',fill_type='solid')

    

def preguntas(hopan):
    """
    

    Parameters
    ----------
    hopan : es un dataframe de pandas.

    Returns
    -------
    ind : (list).
    Regresa una lista con el número de fila en la que detectó cada pregunta.
    Aquí se incluye lo de los número romanos

    """
    
    c=0
    ind=[]
    for i in hopan['Unnamed: 0']:
        a = pd.isna(hopan['Unnamed: 0'][c])
        if a == False:
            ind.append(c)
        c+=1
    bs = ['I)','II)','III)','IV)','V)','VI)','VII)','VIII)']
    ind1 = []+ind
    c = 0
    for elm in ind:
        try:
            nho = hopan.iloc[ind[c]:ind[c+1]]
            
        except:
            nho = hopan.iloc[ind[c]:]
        d = 0
        for i in nho['Unnamed: 2']:
            nter = nho.iat[d,2] 
            nter = str(nter)
            for val in bs:
                if val in nter:
                    ind1.insert(c+1,d+ind[c])
            d += 1
        c +=1
    ind = list(set(ind1))
    ind.sort()
    return ind


def espacio(hopan,preguntas):
    """
    

    Parameters
    ----------
    hopan : dataframe de pandas.
    preguntas : (list). Corresponde a la lista que genera la 
    función preguntas

    Returns
    -------
    di : (dict). Regresa diccionario con el numero de la fila 
    donde hay pregunta como llave, y su valor es el espacio que abarca 
    (es una lista de dos valores)

    """
    
    a = preguntas
    di = {}
    con = 0
    b = len(hopan)
    for i in a:
        try:
            di[con]=[a[con],a[con+1]]
        except:
            di[con]=[a[con],b]
        con+=1
    
    return di



def buscarpalabra(npregunta,palabra,hopan,preguntas):
    """
    

    Parameters
    ----------
    npregunta : (int). Es el número de la pregunta que va a analizar, 
    de la lista de función preguntas 
    palabra : (str). La palabra o frase que se va a 
    buscar (no es búsqueda exacta!, solo que esté en el texto).
    hopan : Dataframe de pandas.
    preguntas : (list).

    Returns
    -------
    entot : (list). Regresa una lista de tuplas con las coordenadas de 
    la palabra

    """
    
    sr = espacio(hopan,preguntas)
    cor = sr[npregunta]
    a = cor[0]
    b = cor[1]
    mdf = hopan.iloc[a:b]
    sa = mdf.fillna('')
    listas = sa.to_numpy().tolist()
    cont = 0
    entot = []
    for lista in listas:
        vo = 0
        try:
            for elm in lista:
                if palabra in elm:
                    sad = (cont,vo)
                    entot.append(sad)
                else:
                    pass
                vo+=1
            cont+=1
        except:
            pass
    return entot



def buscarfrase(npregunta,tuplas,hopan,preguntas):
    """
    

    Parameters
    ----------
    npregunta : (int). Número de pregunta
    tuplas : (list). Lista de tuplas
    hopan : Dataframe de pandas.
    preguntas : (list). Lista de función preguntas

    Returns
    -------
    entot : (list). Regresa lista con las frases que encontró en las 
    coordenadas de la lista de tuplas. Esta es para los catálogos

    """
    
    sr = espacio(hopan,preguntas)
    cor = sr[npregunta]
    a = cor[0]
    b = cor[1]
    mdf = hopan.iloc[a:b]
    entot = []
    for tupla in tuplas:
        frase = mdf.iat[tupla[0],tupla[1]]
        entot.append(frase)
    
    return entot

def saberautosuma(npregunta,hopan,preguntas):
    """
    

    Parameters
    ----------
    npregunta : (int). Número de pregunta
    hopan : Dataframe de pandas.
    preguntas : (list) Lista de preguntas.

    Returns
    -------
    entot : (list). Devuelve lista con coordenadas de la letra "S" de autosuma.
    Es coincidencia exacta

    """
    
    sr = espacio(hopan,preguntas)
    cor = sr[npregunta]
    a = cor[0]
    b = cor[1]
    mdf = hopan.iloc[a:b]
    sa = mdf.isin(['S'])
    listas = sa.to_numpy().tolist()
    cont = 0
    entot = []
    for lista in listas:
        vo = 0
        try:
            for elm in lista:
                if True == elm:
                    sad = (cont,vo)
                    entot.append(sad)
                else:
                    pass
                vo+=1
            cont+=1
        except:
            pass
    return entot

def arrayin(ar,numero,y):
    """
    

    Parameters
    ----------
    ar : array de numpy.
    numero : (int). Número para buscar sus índices en el arreglo
    y : TYPE
        DESCRIPTION.

    Returns
    -------
    res : list. Regresa lista con los indices de los valores buscados

    """
    
    an = np.where(ar==numero)
    indices = an[0]
    res = []
    for ind in indices:
        q = y[ind]
        res.append(q)
    return res



def arreglarcombinado(diccionario):
    """
    

    Parameters
    ----------
    diccionario : (dict). Diccionario con filas y columnas

    Returns
    -------
    dirr : (dict). Diccionario con correcciones para las coordenadas que 
    están convinadas horizontalmente en los encabezados de las tablas

    """
    
    a = diccionario['fila']
    b = diccionario['columna']
    unicos = list(set(a))
    unicos.sort()
    unih = []
    for val in unicos:
        h = a.count(val)
        if h > 2:
            arreglo = np.array(a)
            z = arrayin(arreglo,val,b)
            z1 = max(z)
            if z1 > 10:
                unih.append(val)
    ml = []
    c=0     
    unih.sort()  
    for i in unih[1:]:
        r = i-unih[c]
        try:
            t = unih[c+2]-i
        except:
            t = 0
        if r > 3 and len(ml) ==0:
            ml.append(unih[:c+1])
        if len(ml) !=0 and t > 4:
            q = ml[-1][-1]
            w = unih.index(q)
            ml.append(unih[w+1:c+2])
        if i == unih[-1] and len(ml) !=0:
            q = ml[-1][-1]
            w = unih.index(q)
            ml.append(unih[w+1:])
        if i == unih[-1] and len(ml) ==0:
            ml.append(unih)
        c+=1
    
    for i in ml:
        if len(i) > 1:
            m = max(i)
            if a.count(m+1)==0: #quiere decir que es tabla unica de valores
                arreglo = np.array(a)
                l = arrayin(arreglo,m,b)
                l1 = arrayin(arreglo,m-1,b)
                l2 = arrayin(arreglo,m-2,b)
                l3 = arrayin(arreglo,m-3,b)
                lf = l+l1+l2+l3
                lf = list(set(lf))
                lf.sort()
                for val in lf:
                    if val not in l:
                        ty = a.index(m)
                        a.insert(ty,m)
                        b.insert(ty,val)
                hj = [m-1,m-2,m-3]
                for i in hj:
                    while True:
                        try:
                            arre = np.array(a)
                            nj = np.where(arre==i)
                            nj1= nj[0]
                            a.pop(nj1[0])
                            b.pop(nj1[0])
                        except:
                            break
            if a.count(m+1) > 0: #para todas las demás tablas
                arreglo = np.array(a)
                ver = arrayin(arreglo,m,b)
                pregunta = max(ver) 
                if pregunta > 10: #asegurar que es encabezado y no fila muy larga
                    unipr = []
                    ini = unicos.index(m)
                    cx = 0
                    for i in unicos[ini:]:
                        resta = m + cx - i
                        if resta == 0:
                            unipr.append(i)
                        cx+=1
                    listadelistas = []
                    if len(unipr)<10:
                        for v in unipr[1:]:
                            arreglo = np.array(a)
                            qw = arrayin(arreglo,v,b)
                            listadelistas.append(qw)
                        f1 = max(listadelistas)
                        f2 = max(f1) #define valor máximo de columna en estructura de tabla
                    else:
                        for v in unipr[1:10]:
                            arreglo = np.array(a)
                            qw = arrayin(arreglo,v,b)
                            listadelistas.append(qw)
                        f1 = max(listadelistas)
                        f2 = max(f1)                    
                    l1 = arrayin(arreglo,m-1,b)
                    l2 = arrayin(arreglo,m-2,b)
                    l3 = arrayin(arreglo,m-3,b)
                    lf = ver+l1+l2+l3
                    lf = list(set(lf))
                    lf.sort()
                    for val in lf:
                        if val not in ver and val > f2:
                            ty = a.index(m)
                            a.insert(ty,m)
                            b.insert(ty,val)
                    hj = [m-1,m-2,m-3]
                    for i in hj:
                        while True:
                            try:
                                arre = np.array(a)
                                nj = np.where(arre==i)
                                nj1= nj[0]
                                a.pop(nj1[0])
                                b.pop(nj1[0])
                            except:
                                break
    dirr={'fila':a,'columna':b}
    return dirr   

def imagen(npregunta,hopan,preguntas):
    """
    

    Parameters
    ----------
    npregunta : (int). Número de pregunta
    hopan : Dataframe de pandas.
    preguntas : (list). Lista de preguntas

    Returns
    -------
    ente : (dict).genera un diccionario con filas y columnas en los que se 
    registra un valor de la pregunta

    """
  
    sr = espacio(hopan,preguntas)
    cor = sr[npregunta]
    a = cor[0]
    b = cor[1]
    mdf = hopan.iloc[a:b]
    sa = pd.isna(mdf)
    listas = sa.to_numpy().tolist()
    cont = 0
    entot = {'fila':[],'columna':[]}
    for lista in listas:
        vo = 0
        for elm in lista:
            if False == elm:
                entot['fila'].append(cont)
                entot['columna'].append(vo)          
            else:
                pass
            vo+=1
        cont+=1
    ente = arreglarcombinado(entot)
    return ente    



def coordenadas(dic):
    """
    

    Parameters
    ----------
    dic : (dict). Diccionario con listas de filas y columnas

    Returns
    -------
    lista : (list). Regresa lista de tuplas

    """
    
    x = dic['fila']
    y = dic['columna']
    lista = []
    m = 0
    for i in x:
        c = x[m]
        v = y[m]
        tu = (c,v)
        lista.append(tu)
        m+=1
    return lista




#determinar tipo de pregunta
def analizarcor(npregunta, hopan,preguntas):
    """
    

    Parameters
    ----------
    npregunta : (int). Número de pregunta
    hopan : Dataframe de pandas.
    preguntas : (list). Lista de preguntas

    Returns
    -------
    res : (list). Regresa lista con varias características para determinar, 
    qué tipo de pregunta es, su autosuma y algunas listas con coordenadas 
    para utilizar

    """
    
    dic=imagen(npregunta,hopan,preguntas)
    x = dic['fila']
    # espacios = distancia(x,4)
    y = dic['columna']   
    m1 = max(y)
    #sub = subtotl()
    tabla = 0
    uni = tabunic(x, y)
    tablaunica = 0 #tablas con valor unico, o tablas cortas
    tabla1col = 0
    uno = uncol(x, y)
    delito = buscarpalabra(npregunta, 'Tipo de delito',hopan, preguntas)
    delitotupla = 0
    notabla=[]
    totales = paratotales(npregunta, hopan, preguntas)
    busqueda_condicional = {}
    
    if m1 > 15:
        print('li hzo aquuui',m1)
        if len(delito) > 0 and len(x) > 180:
            tabla = 2 #tabla de delitos
            delitotupla = iral(delito,x,y)
        else:
            tabla = 1 #tabla normal
    if len(uni) > 0:
        unique = list(set(uni))
        tablaunica = unique
    if len(uno) > 0:
        tabla1col = 1
    if tabla == 0:
        notabla= tuplasnt(x,y)
    autosuma = saberautosuma(npregunta,hopan,preguntas)
    autosu = []
    for i in autosuma:
        a = i[0]
        autosu.append(a)
    tabla_partes = continuidad(hopan,npregunta,preguntas)
    if tabla_partes > 0:
        lista_partes = partes(x,tabla_partes)
        lista_partes.pop(0)
        partes_tuplas = tuplas_partes(x, y, lista_partes)
        tabla_partes = partes_tuplas
    if tabla > 0 and totales == 'No':
        tabla = 3 #tabla normal sin totales que validar
    if tabla > 0: 
        busqueda_condicional = varias_busquedas(npregunta,hopan,x,y,tabla_partes,preguntas)
        if tablaunica != 0:
            busqueda_condicional['final'] = 0
            busqueda_condicional['final1'] = 0
    ultimo_escrito=x[-1]
    ultimis = ultimo_escrito + 3 # se hace el ajuste de una vez y es el ultimo valor para poner los mensajes de error en la pregunta
    res = [tabla,tablaunica,tabla1col,uno,
           delitotupla,notabla,autosu,tabla_partes,busqueda_condicional,
           ultimis]
    return res


def depurartu(tuplas,part_tab):
    """
    

    Parameters
    ----------
    tuplas : (list). Lista de tuplas con coordenadas
    part_tab : (list). Lista de tuplas con coordenadas de preguntas por partes

    Returns
    -------
    tuplas : (lista). Lista de tuplas. Remueve las tuplas que se 
    encuentran en otra lista, para no causar confusión al momento de escribir 
    validaciones. 

    """
    for lista in part_tab:
        for tupla in lista:
            if tupla in tuplas:
                tuplas.remove(tupla)
    return tuplas

def varias_busquedas(npregunta,hopan,x,y,part_tab,preguntas):
    """
    

    Parameters
    ----------
    npregunta : (int). Número de pregunta
    hopan : Dataframe de pandas.
    x : (list). Lista de filas
    y : (list). Lista de columnas.
    part_tab : (list). Lista de tuplas
    preguntas : (list). Lista de preguntas

    Returns
    -------
    di : (dict). regresa diccionario con coordenadas de los tipos de 
    busquedade palabras y las tuplas con coordenadas de los encabezados

    """
    
    di = {}
    sinonosabe = buscarpalabra(npregunta, '(1. Sí / 2. No / 9. No se sabe)',
                               hopan,preguntas)
    
    sinon = buscarpalabra(npregunta, '(1. Sí / 2. No / 8. No aplica / 9. No se sabe)',
                               hopan,preguntas)
    noa = buscarpalabra(npregunta, 'No aplica',
                               hopan,preguntas)
    cuatro_numeros = busqueda_4(npregunta,hopan,preguntas)
    tuplas = tuplasnoval(x,y,part_tab)
    final = buscar_final_tabla(x,tuplas)
    di['final1'] = final
    di['final'] = final
    di['tuplas'] = tuplas
    # print('ese es el di de tuplas: ',di)
    cata = buscarpalabra(npregunta, '(ver catálogo)',
                               hopan,preguntas)
    temporal = ['(aaaa)','(mm)','(dd)','Edad\n(años)', '(años)']
    colum_temporales = buscar_varias_cosas(npregunta, hopan, preguntas, temporal)
    
    if len(noa) > 0:
        #version anterior
        # nli = [tupla for tupla in noa if tupla[1] > 3] 
        # if len(sinon)> 0:
        #     ind = []
        #     for tupla in nli:
        #         for tu in sinon:
        #             if tupla[1] != tu[1]:
        #                 ind.append(tupla)
        #     if ind:
        #         ind1 = comsabe(ind,tuplas)
        #         di['NA'] = ind1
        # else:
        #     if nli:
        #         nli1 = comsabe(nli,tuplas)
        #         di['NA'] = nli1
        nli = [tupla for tupla in noa if tupla[1] > 3] 
        if sinon:
            lsinon = [tupla[1] for tupla in sinon]
            ind = []
            for tupla in nli:
                if tupla[1] not in lsinon:
                    ind.append(tupla)
                # for tu in sinon:
                #     if tupla[1] != tu[1]:
                #         ind.append(tupla)
            if ind:
                final = buscar_final_tabla(x,ind)
                di['final'] = final
                ind1 = comsabe(ind,tuplas)
                di['NA'] = ind1
        else:
            if nli:
                final = buscar_final_tabla(x,nli)
                di['final'] = final
                nli1 = comsabe(nli,tuplas)
                di['NA'] = nli1
    
    if len(sinonosabe) > 0:
        final = buscar_final_tabla(x,sinonosabe)
        di['final'] = final
        sabe = comsabe(sinonosabe,tuplas) #arreglar fila de la tupla por celdas combinadas
        di['sabe'] = sabe
    
    if len(cuatro_numeros) > 0:
        final = buscar_final_tabla(x,cuatro_numeros)
        di['final'] = final
        sabe1 = comsabe(cuatro_numeros,tuplas) #arreglar fila de la tupla por celdas combinadas
        di['sabe1'] = sabe1  
    
    if len(cata) > 0:
        
        catalogos = {}
        frases = buscarfrase(npregunta,cata,hopan,preguntas)
        catalogos['frases'] = frases
        cor = comsabe(cata,tuplas)
        catalogos['coordenadas'] = cor
        di['catalogos'] = catalogos      
    
    if len(sinon) > 0:
        final = buscar_final_tabla(x,sinon)
        di['final'] = final
        sinon1 = comsabe(sinon, tuplas)
        di['noaplica'] = sinon1
    
    if colum_temporales:
        
        if '(años)' in colum_temporales:
            final = buscar_final_tabla(x,colum_temporales['(años)'])
            di['final'] = final
        if '(aaaa)' in colum_temporales:
            final = buscar_final_tabla(x,colum_temporales['(aaaa)'])
            di['final'] = final
        
        nd = {}
        for key in colum_temporales:
            
            lista = comsabe(colum_temporales[key],tuplas)
            nd[key] = lista
        if 'Edad\n(años)' in nd and '(años)' in nd:
            for tupla in nd['(años)']:
                if tupla in nd['Edad\n(años)']:
                    nd['(años)'].remove(tupla)
        
        di['temporales'] = nd
    
    return di

def buscar_final_tabla(x,tuplas):
    """
    

    Parameters
    ----------
    x : list. Lista de filas
    tuplas : list. Lista de tuplas de la pregunta

    Returns
    -------
    None.

    """
    fin = distancia(x,1)
    inicio = tuplas[0][0]
    p = []
    for i in fin:
        fila = x[i]
        if fila>inicio:
            r = fila-inicio
            p.append(r)
    if not p:
        p = [0]
    return p[0]

def busqueda_4(npregunta,hopan,preguntas):
    """
    

    Parameters
    ----------
    npregunta : TYPE
        DESCRIPTION.
    hopan : TYPE
        DESCRIPTION.
    preguntas : TYPE
        DESCRIPTION.
    

    Returns
    -------
    a : (list). Regresa una lista con tuplas coordenadas de donde 
    encontró la frase

    """
    busquedas = ['(1. Sí / 2. En proceso de certificación / 3. No / 9. No se sabe)',
                 '(1. Sí / 2. En proceso de integración / 3. No / 9. No se sabe)',
                 '(1. Voz / 2. Voto / 3. Voz y voto / 9. No se sabe)']
    #solo puede haber una en la pregunta!!
    for busqueda in busquedas:
        a = buscarpalabra(npregunta, busqueda, hopan,preguntas)
        if len(a) > 0:
            break
    return a

def buscar_varias_cosas(npregunta,hopan,preguntas,busquedas):
    """
    

    Parameters
    ----------
    npregunta : TYPE
        DESCRIPTION.
    hopan : TYPE
        DESCRIPTION.
    preguntas : TYPE
        DESCRIPTION.
    busquedas : list. Lista de strings con las cosas a buscar

    Returns
    -------
    a : (list). Regresa una lista con tuplas coordenadas de donde 
    encontró la frase

    """
    
    r = {}
    for busqueda in busquedas:
        a = buscarpalabra(npregunta, busqueda, hopan,preguntas)
        if len(a) > 0:
            r[busqueda] = a
    return r


def comsabe(tuplasa, tuplasb):
    """
    

    Parameters
    ----------
    tuplasa : (list). Lista de tuplas a arreglar
    tuplasb : (list). Lista de tuplas que están bien.

    Returns
    -------
    res : (list). lista de tuplas a, pero arreglada. Esta funcion es 
    necesaria debido a las diferencias que puede haber en las cordenadas por
    celdas combinadas en horizontal

    """
    comprobar = [tupla[0] for tupla in tuplasa]
    unic = list(set(comprobar))
    diferencias = distancia(unic, 4)
    
    resta = unic[0] - tuplasb[0][0]
    if resta>5:
        return tuplasa
    else:
        if diferencias:
            tabla1 = [tupla for tupla in tuplasa if tupla[0] in unic[:diferencias[0]+1]] 
            tabla2omas = [tupla for tupla in tuplasa if tupla[0] not in unic[:diferencias[0]+1]]
            res = [(tuplasb[0][0],tupla[1]) for tupla in tabla1]
            res += tabla2omas
        else:
            res = [(tuplasb[0][0],tupla[1]) for tupla in tuplasa]
        return res

def tuplasnoval(x,y,part_tab):
    """
    

    Parameters
    ----------
    x : (list). Lista de filas.
    y : (list). Lista de columnas.
    part_tab : (list). Lista de tuplas.

    Returns
    -------
    ndefi : (list). Lista de tuplas.

    """
    unicos = list(set(x))
    posibles = []
    for val in unicos:
        a = x.count(val)
        if a > 1:
            posibles.append(val)
    tuplas = []
    c = 0
    for co in x:
        if co in posibles:
            tup = (co,y[c])
            tuplas.append(tup)
        c += 1
    li =[]
    for tupla in tuplas:
        if tupla[1] > 15 and tupla[0] < 25 and tupla[1]<31:
            li.append(tupla)
    li1 = []
    for tupla in li:
        a = tupla[0]
        li1.append(a)
    li2 = list(set(li1))
    defi = []
    #Chek
    # for tu in tuplas:
    #     if tu[0] in li2:
    #         defi.append(tu)
    #experimental
    li2 = reducir(li2)
    for tu in tuplas:
        if tu[0] in li2:
            defi.append(tu)

    # unificar filas por celdas combinadas
    if part_tab != 0:
        filtro_partab = depurartu(defi,part_tab)
        ndefi = comprobaciontuplas(filtro_partab)
    else:
        ndefi = comprobaciontuplas(defi)
    return ndefi

def reducir(lista):
    """reduce lista"""
    if len(lista)>1:
        
        dist = distancia(lista,1)
        if len(dist)>0:
            limite = dist[0]+1
            rlista = lista[:limite]
            nlista = rlista
        else:
            nlista = lista
        
    else:
        nlista = lista
    return nlista

def comprobaciontuplas(tuplas):
    """
    

    Parameters
    ----------
    tuplas : (list). Lista de tuplas.

    Returns
    -------
    res : (list). Lista de tuplas ajustadas al valor mayor de la fila.

    """
    filas = [tupla[0] for tupla in tuplas]
    unicos = list(set(filas))
    m = max(unicos)
    res = [(m,tupla[1]) for tupla in tuplas]
    return res

def paratotales(npregunta,hopan,preguntas):
    """
    

    Parameters
    ----------
    npregunta : TYPE
        DESCRIPTION.
    hopan : TYPE
        DESCRIPTION.
    preguntas : TYPE
        DESCRIPTION.

    Returns
    -------
    res : (dict). Regresa diccionario con listas de tuplas sobre coordenadas
    en donde hay totales y/o subtotales.
    
    En caso de que no haya, regresa 'No'
    """
    
    sr = espacio(hopan,preguntas)
    cor = sr[npregunta]
    a = cor[0]
    b = cor[1]
    mdf = hopan.iloc[a:b]
    total = 'Total' in mdf.values
    subtotal = 'Subotal' in mdf.values
    
    if total == True or subtotal == True:
        t = totorsubtot(mdf,'Total')
        s = totorsubtot(mdf,'Subtotal')
        print('van total y subtotal!: ',t,s)
        res ={'total':t,'subtotal':s}
    else:
        return 'No'
    return res
    

def totorsubtot(dataframe,vaal):
    """
    

    Parameters
    ----------
    dataframe : TYPE
        DESCRIPTION.
    vaal : (str). Palabra a buscar

    Returns
    -------
    res : (list). genera lista de tuplas con coordenadas del vaal asignado;
    Es coincidencia exacta

    """
    
    sa = dataframe.isin([vaal])
    listas = sa.to_numpy().tolist()
    cont = 0
    entot = {'fila':[],'columna':[]}
    for lista in listas:
        vo = 0
        for elm in lista:
            if True == elm:
                entot['fila'].append(cont)
                entot['columna'].append(vo)          
            else:
                pass
            vo+=1
        cont+=1
    res = tuplasnt(entot['fila'],entot['columna'])
    return res  

def tuplas_partes(x,y,lista):
    "Recibe listas: filas, columnas y numeros de partes para hacer las tuplas, regresa lista de listas de tuplas"
    a = tuplasnt(x, y)
    res = []
    for i in lista:
        l = []
        for tupla in a:
            if tupla[0] == i:
                l.append(tupla)
        l.sort()
        res.append(l)
                
    return res
    
def partes(lista_filas,numero):
    "recibe lista y regresa otra lista con los numeros más repetidos de acuerdo al valor ingresado en numero "
    l = Counter(lista_filas)
    r1 = l.most_common(numero)
    r = []
    for val in r1:
        a = val[0]
        r.append(a)
    r.sort()
    return r

def continuidad(hopan,npregunta,preguntas):
    "Recibe dataframe y numero de pregunta, regresa un número si es que hay continuidad en las tablas, el numero indica cuantas tablas son continuas, si es cero, pues ninguna"
    cor_pal = buscarpalabra(npregunta, '(1 de ', hopan,preguntas)
    print('continuidad, cor_pal: ', cor_pal)
    pro = preguntas        
    pregunta = pro[npregunta]
    if len(cor_pal) > 0:
        c1 = cor_pal[0]
        c1 = list(c1)
        fila = pregunta+c1[0]
        ter = hopan.iat[fila,c1[1]]
        print('continuidad, ter: ', cor_pal)
        try:
            ni = ter[-2]
            ni = int(ni)
        except:
            ni = 0
    else:
        ni = 0
    print('ese fue ni, continuidsd: ', ni)
    return ni


def tuplasnt(x,y):
    """
    

    Parameters
    ----------
    x : (list). Lista filas
    y : (list). Lista columnas

    Returns
    -------
    tuplas : (list). Lista de tuplas con coordenada sa partir de X y Y

    """
    tuplas=[]
    ind=0
    for i in x:
        a = (i,y[ind])
        tuplas.append(a)
        ind+=1
    return tuplas

def iral(tuplas,x,y):
    "Regresa las tuplas donde se ubican las columnas en preguntas de tipo de delito"
    tres =[]
    
    for tupla in tuplas:
        
        ind = 0
        for val in x:
 
            if tupla[0] == val:
                
                y1 = y[ind]
                if y1 > tupla[1]:
                    print('int 1.1')
                    tres.append((val,y1))
            if tupla[0]+1 == val:
                
                y1 = y[ind]
                if y1 > tupla[1]:
                    print('int 2.1')
                    tres.append((val,y1))
            if tupla[0]+2 == val:
                
                y1 = y[ind]
                if y1 > tupla[1]:
                    tres.append((val,y1))
            
            if tupla[0]+3 == val:
                
                y1 = y[ind]
                if y1 > tupla[1]:
                    tres.append((val,y1))
            ind+=1
    
    eli=[]
    for i in tres:
        e = i[0]
        eli.append(e)
    ti = []
    
    for i in eli:
        r = eli.count(i)
        if r > 2:
            ti.append(i)
    if not ti:
        ti.append(eli[0])
    fi = []
    fri = list(set(ti))
    for i in fri:
        for val in tres:       
            if i in val:
                fi.append(val)

    siss = {'fila':x,'columna':y}            
    cors = coordenadas(siss)
    desa = []  
    tom = 0
    for tupla in tuplas:
        re = tupla[1]+1
        inx = fri[tom]
        dd = []
        for i in range(inx,inx+167):
            wa = (i,re)
            if wa in cors:
                dd.append(i)
        desa.append(dd)
        tom+=1
    xn = []
    for tupla in fi:
        a = tupla[0]
        xn.append(a)
    xn1 = list(set(xn))
    pre = []
    m = 0
    for v in xn1:
        try:
            a = xn1[m+1]-v
            if a > 165:
                pre.append(v)
        except:
            a = xn1[m] - xn1[m-1] 
            if a > 165:
                pre.append(v)
        m+=1
    if len(pre) == 0:
        pre.append(xn1[-1])
    defi = []
    for val in pre:
        for tupla in fi:
            if val == tupla[0]:
                defi.append(tupla)
    
    respu = {'tuplas':defi,'desagregados':desa}
    return respu



def uncol(x,y):
    """
    

    Parameters
    ----------
    x : (list). Lista de filas
    y : (list). Lista de columnas.

    Returns
    -------
    resul : (list). Lista de tuplas para preguntas de una columna

    """
    ex = []
    tu =[]
    c = 0
    resul = []
    for i in x:
        a = x.count(i)
        if a == 2:
            y1 = y[c]
            r = (i,y1)
            tu.append(r)
            ex.append(i)
        c+=1

    ex1 = list(set(ex))
    for i in ex1:
        resta = []
        for tupla in tu:
            if i in tupla:
                resta.append(tupla[1])
        ta = resta[1] - resta [0]
        if ta > 15:
            resul.append((i,resta[1]))
    
    return resul



def tabunic(x,y):
    """
    

    Parameters
    ----------
    x : list. Lista de filas
    y : list. Lista de columnas.

    Returns
    -------
    res : regresa una lista con el numero de la fila donde se identificó que existe una tabla de fila única de respuesta.

    """
    
    c = 0
    o = []
    res = []
    s = distancia(x, 2)
    tx =[]
    for i in y:
        if i > 9:
            o.append(c)
        c+=1
    for i in s:
        g =x[i]
        tx.append(g)
    for i in o:
        h = x[i]
        a = x.count(h)
        if a > 1:
            w = x[i]            
            if w in tx:
                posx=[]
                so = 0
                for ti in x:
                    if ti==h:
                        posx.append(so)
                    so+=1
                for val in posx: 
                    ft = (h,y[val])
                    res.append(ft)
    comprobar = [val[0] for val in res]
    try:
        if min(comprobar) > 15:
            res = []  
    except:
        pass
    return res
            


def distancia(lista,nfilas):
    """
    

    Parameters
    ----------
    lista : list. Lista con numeros
    nfilas : int. La distancia o diferencia a medir

    Returns
    -------
    li : list. Regresa una lista con los indices de la lista de entrada en donde se cumple la diferencia señalada 

    """
    
    li = []
    con = 0
    for i in lista:
        try:
            a = lista[con+1]-lista[con]
        except:
            a = 0
        if a > nfilas:
            li.append(con)
        con+=1
    return li




#Conseguir coordenadas de respuesta y enlazarlas a variables


def getcor(npregunta,hopan,preguntas):
    """
    

    Parameters
    ----------
    npregunta : int. El numero de la pregunta en la que se va a hacer el analisis dentro del dataframe de pandas
    hopan : Dataframe de pandas.
    preguntas : Lista de preguntas

    Returns
    -------
    res : dict. Devuelve un diccionario con fila de inicio, fila de fin y coordenadas de columnas de la pregunta

    """
    
    dic = imagen(npregunta,hopan,preguntas)
    x = dic['fila']
    y = dic['columna']
    # dis = distancia(x,4)
    pfin = distancia (x,1)
    pf = [x[i] for i in pfin]
    res = {'final':[],'inicio':[],'coord':[]}
    # for i in dis:
    #     res['final'].append(x[i]) 
    mayores9= may9(y)
    i1erf = []         #primer filtro de inicio, regresa indices para listas de dic
    for i in mayores9:
        y1 = y[i]
        x1 = x[i]
        tu = (x1+1,y1)
        liscor = coordenadas(dic)
        try:
            liscor.index(tu)
        except:
            res['coord'].append(tu)
            i1erf.append(x1+1)
    se = pd.Series(i1erf)
    se1 = se.to_frame()
    duplicate = se1[se1.duplicated(keep = 'last')]
    sas = duplicate[0].unique()
    sas1 = sas.tolist()
    for i in sas1:       
        res['inicio'].append(i)
    nuevas_tuplas = []
    for tupla in res['coord']:
        if tupla[0] in res['inicio']:
            nuevas_tuplas.append(tupla)
    res['coord'] = nuevas_tuplas
    chek = [i for i in pf if i > res['inicio'][0]]
    try:
        res['final'].append(chek[0])
    except:
        pass
    return res




def distanciaigual(lista):
    "medir distancia entre filas"
    li = []
    con = 0
    for i in lista:
        try:
            a = lista[con+1]-lista[con]
        except:
            a = 1
        if a == 0:
            li.append(con)
        con+=1
    return li

def may9(lista):
    "Recibe lista y regresa otra con los elementos que se son mayores de 2"
    res = []
    co = 0
    for i in lista:
        if i > 2:
            res.append(co)
        co+=1
    return res

def p_especificas(hopan,hoja,preguntas_validar):         
    """
    hopan: dataframe de pandas
    hoja: hoja de excel de openpyxl
    preguntas: list. Lista con preguntas exactamente como están en la 
    columna A del excel
    
    
    
    """
    
    
    pro = preguntas(hopan) 
    t = espacio(hopan,pro)
    con = 0
    for i in t:
               
        pregunta = pro[con] #aqui va fila inicio de la pregunta para iterar
        nter = hopan.iat[pregunta,0] 
        
        if nter in preguntas_validar:
            r = analizarcor(con,hopan,pro)
            totales = paratotales(con, hopan,pro)
            part_tab = r[7]
            part = r[7] #se hace copia porque la variable se modifica y a las condicionales de abajo ya no llega bien
            print('este es r de pregunta ', nter, ':', r)
            letras_validadas = []
            # alfa = clasificador(r,totales)
            # print('el alfa: ', alfa)
            
            if 'sabe' in r[8] or 'sabe1' in r[8] or 'noaplica' in r[8] or 'NA' in r[8]:
                diccio = r[8]
                tuplas = diccio['tuplas']
                ad = masdeunatablauni(tuplas)
                
                for i in ad:
                    if nter != '8.1.-':
                        freal = pregunta+i+1
                        validarcondicional(diccio,part,freal,i,hoja)
                    else: #porque pusieron un encabezado todo combinado y había que sumar un uno para ajustar, ya que prolema de merge cell
                        
                        freal = pregunta+i+1+1
                        for tuplas in part:
                            npart = [[(tupla[0]-1,tupla[1]) for tupla in tuplas]] # se hace ajuste y es exclusivo para pregunta 8.1
                        validarcondicional(diccio,npart,freal,i,hoja)
            if 'temporales' in r[8]:
                # diccionario = r[8]['temporales']
                validar_temporal(pregunta, hoja, r[8])
            
            if r[0]==1 and r[1]==0: #si es tabla general
                print('incorrecto')
                a = getcor(con,hopan,pro)
                inicios = a['inicio']
                tupla = a['coord']
                tupla.sort()
                print('inicios para tabla normal: ', a)
                c=0
                for i in inicios:
                    freal = pregunta+i #fila real
                    fin = r[6]
                    ntuplas = []
                    for tupl in tupla:
                        tfil = tupl[0]
                        if tfil == i:
                            ntuplas.append(tupl)
                    tuplas = subtuplas(totales, ntuplas,part_tab,r[8])
                    # print ('salida subtuplas: ', len(tuplas), len(ntuplas),ntuplas,i)
                    if fin == []: #para ajustar perguntas de tabla que no tiene autosuma 
                        fin =[0 - pregunta for i in range(0,len(inicios))]
                    if len(tuplas) == len(ntuplas):
                        ldeletras= letraslide(c)
                        if fin == []:
                            hay = variables(i,tuplas,freal,0,hoja,part_tab,ldeletras,'si')
                        else:
                            hay = variables(i,tuplas,freal,fin[c]+pregunta,hoja,part_tab,ldeletras,'si') 
                        hay = [hay[0]]
                    if c == len(fin):
                        break
                    if len(tuplas) != len(ntuplas):
                        part_tab = 0
                        c1 = 0
                        extraer_letras=[]
                        for lista in tuplas:
                            print('entro aqui 1')
                            ldeletras= letraslide(c1)
                            
                            # if lista ==tuplas[-1] and len(totales['subtotal']) > 0 and lista[-1] != 'a':
                            #     print('entro aqui 2')
                            #     hay = validarTS(i,lista,freal,fin[c]+pregunta,hoja,ldeletras)
                               
                            if lista[-1] == 'a':
                                lista.remove('a')
                                print('entro aqui 3')
                                hay = validarTS(i,lista,freal,fin[c]+pregunta,hoja,ldeletras)
                            else:
                                print('entro aqui 4')
                                hay = variables(i,lista,freal,fin[c]+pregunta,hoja,part_tab,ldeletras,'no') 
                                hay = hay[0]
                            letras_validadas.append(hay)    
                            extraer_letras.append(ldeletras[:6])
                            c1+=1
                        
                    if fin[c]+pregunta == 0:
                        menerrorsub(freal+2, a['final'][0]+2+pregunta, hoja, extraer_letras,r[9]+pregunta)
                    if fin[c]+pregunta != 0:
                        menerrorsub(freal+2, fin[c]+pregunta+2, hoja, extraer_letras,r[9]+pregunta)
                    c+=1
            if r[1]!=0  and r[0]==1: #tablas de filas unicas
                ad = masdeunatablauni(r[1])
                print('es correcto')
                for i in ad:
                    
                    freal = pregunta+i+1 #mas uno porque es la fila del titulo de columnas
                    r[1].sort()
                    print('akkkkk')
                    tuplas = subtuplas(totales, r[1],part_tab,r[8])
                    ntuplas = []
                    for tupl in r[1]:
                        tfil = tupl[0]
                        if tfil == i:
                            ntuplas.append(tupl)
                    if len(tuplas) == len(r[1]):
                        ldeletras= letraslide(0)
                        hay = variables(i,ntuplas,freal,0,hoja,part_tab,ldeletras,'si')
                        hay = hay[0]
                    
                    else:
                        extraer_letras=[]
                        part_tab = 0
                        c = 0
                        for lista in tuplas:
                            ldeletras= letraslide(c)
                            
                            # if lista ==tuplas[-1] and len(totales['subtotal']) > 0:
                            #     hay = validarTS(i,lista,freal,0,hoja,ldeletras)
                            
                            if lista[-1] == 'a':
                                lista.remove('a')
                                hay = validarTS(i,lista,freal,0,hoja,ldeletras)
                            
                            else:
                                hay = variables(i,lista,freal,0,hoja,part_tab,ldeletras,'no')
                                hay = hay[0]
                            letras_validadas.append(hay)
                            extraer_letras.append(ldeletras[:6])
                            c+=1
                    menerrorsub(freal+2, freal+2, hoja, extraer_letras,r[9]+pregunta)
            
            if r[0]==2:
                p = r[4]
                u = p['desagregados']
                ad = masdeunatablauni1(p['tuplas'])
                delito = buscarpalabra(con, 'Tipo de delito',hopan,pro)
                k = 0
                
                for i in ad:
                    tp = p['tuplas']
                    tp.sort()
                    tuplas = subtuplas(totales, tp, part_tab,r[8])
                    codelitos = u[k]
                    de = delito[k]
                    freal = i+1+pregunta
                    fin = r[6]
                    if len(tuplas) == len(tp):
                        ldeletras= letraslide(0)
                        hay = vardelitos(i+1, p['tuplas'], freal,de,codelitos,
                                   pregunta+2,fin[k]+pregunta,hoja,part_tab,ldeletras,'si')
                        letras_validadas.append(hay)
                    else:
                        extraer_letras=[]
                        part_tab = 0
                        c = 0
                        for lista in tuplas:
                            ldeletras= letraslide(c)
                            
                            # if lista ==tuplas[-1] and len(totales['subtotal']) > 0:
                            #     hay = validarTS(i,lista,freal,0,hoja,ldeletras)
                            
                            if k == len(fin):
                                break
                            
                            if lista[-1] == 'a':
                                lista.remove('a')
                                hay = validarTSD(i,lista,freal,fin[k]+pregunta,hoja,ldeletras,pregunta+2,codelitos)
                                letras_validadas.append(hay)
                            else:
                                hay = vardelitos(i+1, lista, freal,de,codelitos,
                                   pregunta+2,fin[k]+pregunta+1,hoja,part_tab,ldeletras,'no')
                                letras_validadas.append(hay)
                            extraer_letras.append(ldeletras[:6])
                            c+=1
                    
                    menerrorsub(freal+2, fin[k]+pregunta+2, hoja, extraer_letras,r[9]+pregunta)
                    k+=1
                na_desagregados = x_valida.juntar()
                gris_desagregados(na_desagregados,letras_validadas,hoja)
            if r[2]==1 and r[0]==1: #tablas de unica columna de respuesta
                ad = masdeunatablauni(r[3])
                c = 0
                ad.sort()
                ldeletras= letraslide(0)
                for i in ad:             
                    freal = pregunta+i+1
                    fin = r[6]
                    variablesl1(i, r[3], freal,fin[c]+pregunta,hoja,ldeletras)
                    c+=1
                    
                if len(r[6]) > len(ad):
                    print('Tabla con columnas de respuesta unica no detectada')
            
            if r[0] == 0: #Las que no son tablas
                tuplas = r[5] #esto debe ser lista de tuplas
                corde = clasi(tuplas)
                freal =  pregunta+2
                variablesnotab(corde,freal,hoja,r[9])
            
            
                    
            if r[0] == 3:
                diccio = r[8]
                tuplas = diccio['tuplas']
                ad = masdeunatablauni(tuplas)
                fin = r[6]
              
                if fin ==[]:
                    fin = [0 for i in ad]
          
                c = 0
                for i in ad:
                    freal = pregunta+i+1
                    if c == len(fin):
                        break
                    soloautosuma(diccio,part_tab,freal,i,hoja,fin[c])
                    c += 1
            if 'catalogos' in r[8]:
                di = r[8]
                validar_catalogo(pregunta,di,hoja)
            
    
            if totales != 'No':
                #if len(totales['total'])==1 and len(totales['subtotal'])>0: #para poner el gris del na a toda la fila
                fn = final_tabla.juntar()
                if not fn:
                    fn =[0]
                poner_gris(letras_validadas,hoja,fn[0])
            
            print('Todo bien con pregunta ', nter)    
            con+=1
        else:
            con+=1
            pass
        
     
    return 
errores = []

def procesarcoor(hopan, hoja, seccion):         #funcion principal!!
    """
    hopan: dataframe de pandas
    hoja: hoja de excel de openpyxl
    
    
    """
    
    
    pro = preguntas(hopan) 
    t = espacio(hopan,pro)
    con = 0
    for i in t:
               
        pregunta = pro[con] #aqui va fila inicio de la pregunta para iterar
        nter = hopan.iat[pregunta,0]
        try:
            frame.agregar_valor_acolumna(nter,'pregunta')
            frame.agregar_valor_acolumna(seccion,'seccion')
            r = analizarcor(con,hopan,pro)
            totales = paratotales(con, hopan,pro)
            part_tab = r[7]
            part = r[7] #se hace copia porque la variable se modifica y a las condicionales de abajo ya no llega bien
            print('este es r de pregunta ', nter, ':', r)
            letras_validadas = []
            # alfa = clasificador(r,totales)
            # print('el alfa: ', alfa)
            
            if 'sabe' in r[8] or 'sabe1' in r[8] or 'noaplica' in r[8] or 'NA' in r[8]:
                diccio = r[8]
                tuplas = diccio['tuplas']
                ad = masdeunatablauni(tuplas)
                
                for i in ad:
                    if nter != '8.1.-':
                        freal = pregunta+i+1
                        validarcondicional(diccio,part,freal,i,hoja)
                    else: #porque pusieron un encabezado todo combinado y había que sumar un uno para ajustar, ya que prolema de merge cell
                        
                        freal = pregunta+i+1+1
                        for tuplas in part:
                            npart = [[(tupla[0]-1,tupla[1]) for tupla in tuplas]] # se hace ajuste y es exclusivo para pregunta 8.1
                        validarcondicional(diccio,npart,freal,i,hoja)
            if 'temporales' in r[8]:
                # diccionario = r[8]['temporales']
                validar_temporal(pregunta, hoja, r[8])
            
            if r[0]==1 and r[1]==0: #si es tabla general
                a = getcor(con,hopan,pro)
                inicios = a['inicio']
                tupla = a['coord']
                tupla.sort()
                print('inicios para tabla normal: ', a)
                c=0
                for i in inicios[:1]:
                    freal = pregunta+i #fila real
                    fin = r[6]
                    ntuplas = []
                    for tupl in tupla:
                        tfil = tupl[0]
                        if tfil == i:
                            ntuplas.append(tupl)
                    tuplas = subtuplas(totales, ntuplas,part_tab,r[8])
                    # print ('salida subtuplas: ', len(tuplas), len(ntuplas),ntuplas,i)
                    if fin == []: #para ajustar perguntas de tabla que no tiene autosuma 
                        fin =[0 - pregunta for i in range(0,len(inicios))]
                    if len(tuplas) == len(ntuplas):
                        ldeletras= letraslide(c)
                        if fin == []:
                            hay = variables(i,tuplas,freal,0,hoja,part_tab,ldeletras,'si')
                        else:
                            hay = variables(i,tuplas,freal,fin[c]+pregunta,hoja,part_tab,ldeletras,'si') 
                        hay = [hay[0]]
                    if c == len(fin):
                        break
                    if len(tuplas) != len(ntuplas):
                        part_tab = 0
                        c1 = 0
                        extraer_letras=[]
                        tuplas = [lista for lista in tuplas if lista] #limpieza de listas vacias
                        print('revisar', tuplas,ntuplas)
                        for lista in tuplas:
                            print('entro aqui 1')
                            ldeletras= letraslide(c1)
                            
                            # if lista ==tuplas[-1] and len(totales['subtotal']) > 0 and lista[-1] != 'a':
                            #     print('entro aqui 2')
                            #     hay = validarTS(i,lista,freal,fin[c]+pregunta,hoja,ldeletras)

                            if lista[-1] == 'a':
                                lista.remove('a')
                                print('entro aqui 3')
                                hay = validarTS(i,lista,freal,fin[c]+pregunta,hoja,ldeletras)
                            else:
                                print('entro aqui 4')
                                hay = variables(i,lista,freal,fin[c]+pregunta,hoja,part_tab,ldeletras,'no') 
                                hay = hay[0]
                            letras_validadas.append(hay)    
                            extraer_letras.append(ldeletras[:6])
                            c1+=1
                        
                    if fin[c]+pregunta == 0:
                        try:
                            menerrorsub(freal+2, a['final'][0]+2+pregunta, hoja, extraer_letras,r[9]+pregunta)
                        except:
                             menerrorsub(freal+2, 7+2+pregunta, hoja, extraer_letras,r[9]+pregunta) #correccion especial para pregunta 1.54 de CNIJF 2022
                    if fin[c]+pregunta != 0:
                        menerrorsub(freal+2, fin[c]+pregunta+2, hoja, extraer_letras,r[9]+pregunta)
                    c+=1
            if r[1]!=0  and r[0]==1: #tablas de filas unicas
                print('correcto')
                ad = masdeunatablauni(r[1])
                for i in ad[0:1]:
                    
                    freal = pregunta+i+1 #mas uno porque es la fila del titulo de columnas
                    r[1].sort()
                    tup15 = r[8]['tuplas']
                    tup15.sort()
                    if nter == '2.16.-':
                        npart = []
                        for bi in part_tab:
                            np = []
                            for tupla in bi:
                                a = tupla[0]-1 
                                np.append((a,tupla[1]))
                            npart.append(np)
                        tuplas = subtuplas(totales,tup15 ,npart,r[8])
                    
                    else:
                        r[8]['tuplas'] = r[1]
                        tuplas = subtuplas(totales,tup15 ,part_tab,r[8])
                    ntuplas = []
                    for tupl in r[1]:
                        tfil = tupl[0]
                        if tfil == i:
                            ntuplas.append(tupl)
                    if len(tuplas) == len(r[1]):
                        ldeletras= letraslide(0)
                        print('4444')
                        hay = variables(i,ntuplas,freal,0,hoja,part_tab,ldeletras,'si')
                        hay = hay[0]
                        
                    else:
                        extraer_letras=[]
                        part_tab = 0
                        c = 0
                        for lista in tuplas:
                            ldeletras= letraslide(c)
                            
                            # if lista ==tuplas[-1] and len(totales['subtotal']) > 0:
                            #     hay = validarTS(i,lista,freal,0,hoja,ldeletras)
                            
                            if lista[-1] == 'a':
                                lista.remove('a')
                                hay = validarTS(i,lista,freal,0,hoja,ldeletras)
                            
                            else:
                                hay = variables(i,lista,freal,0,hoja,part_tab,ldeletras,'no')
                                hay = hay[0]
                            letras_validadas.append(hay)
                            extraer_letras.append(ldeletras[:6])
                            c+=1
                            
                    menerrorsub(freal+2, freal+2, hoja, extraer_letras,r[9]+pregunta)
            
            if r[0]==2:
                p = r[4]
                u = p['desagregados']
                ad = masdeunatablauni1(p['tuplas'])
                delito = buscarpalabra(con, 'Tipo de delito',hopan,pro)
                k = 0
                
                for i in ad:
                    tp = p['tuplas']
                    tp.sort()
                    tuplas = subtuplas(totales, tp, part_tab,r[8])
                    codelitos = u[k]
                    de = delito[k]
                    freal = i+1+pregunta
                    fin = r[6]
                    if len(tuplas) == len(tp):
                        ldeletras= letraslide(0)
                        hay = vardelitos(i+1, p['tuplas'], freal,de,codelitos,
                                   pregunta+2,fin[k]+pregunta,hoja,part_tab,ldeletras,'si')
                        letras_validadas.append(hay)
                    else:
                        extraer_letras=[]
                        part_tab = 0
                        c = 0
                       
                        for lista in tuplas:
    
                            ldeletras= letraslide(c)
                            # if lista ==tuplas[-1] and len(totales['subtotal']) > 0:
                            #     hay = validarTS(i,lista,freal,0,hoja,ldeletras)
                            
                            if k == len(fin):
                                break
                            
                            if lista[-1] == 'a':
                                lista.remove('a')
                                hay = validarTSD(i,lista,freal,fin[k]+pregunta,hoja,ldeletras,pregunta+2,codelitos)
                                letras_validadas.append(hay)
                            else:
                                hay = vardelitos(i+1, lista, freal,de,codelitos,
                                   pregunta+2,fin[k]+pregunta+1,hoja,part_tab,ldeletras,'no')
                                letras_validadas.append(hay)
                            extraer_letras.append(ldeletras[:6])
                            c+=1
                    
                    menerrorsub(freal+2, fin[k]+pregunta+2, hoja, extraer_letras,r[9]+pregunta)
                    k+=1
                na_desagregados = x_valida.juntar()
                print('letras validadas,desag',na_desagregados,letras_validadas)
                gris_desagregados(na_desagregados,letras_validadas,hoja)
            if r[2]==1 and r[0]==1 and r[1]==0: #tablas de unica columna de respuesta
                ad = masdeunatablauni(r[3])
                c = 0
                ad.sort()
                ldeletras= letraslide(0)
                for i in ad:             
                    freal = pregunta+i+1
                    fin = r[6]
                    variablesl1(i, r[3], freal,fin[c]+pregunta,hoja,ldeletras)
                    c+=1
                    
                if len(r[6]) > len(ad):
                    print('Tabla con columnas de respuesta unica no detectada')
            
            if r[0] == 0: #Las que no son tablas
                tuplas = r[5] #esto debe ser lista de tuplas
                corde = clasi(tuplas)
                freal =  pregunta+2
                variablesnotab(corde,freal,hoja,r[9])
            
            
                    
            if r[0] == 3:
                diccio = r[8]
                tuplas = diccio['tuplas']
                if r[3]:
                    diccio['tuplas'] = r[3]
                ad = masdeunatablauni(tuplas)
                fin = r[6]
              
                if fin ==[]:
                    fin = [0 for i in ad]
          
                c = 0
                for i in ad:
                    freal = pregunta+i+1
                    if c == len(fin):
                        break
                    soloautosuma(diccio,part_tab,freal,i,hoja,fin[c])
                    c += 1
                if r[4]: #si es de delitos meter autosuma de delitos
                    tupla = r[4]['tuplas'] #solo tendrá una tupla
                    abc1 = list(string.ascii_uppercase)+['AA','AB','AC','AD']
                    columna = [abc1[tupla[0][1]]]
                    autosumadelito(columna, pregunta+1, hoja, tupla[0])
            if 'catalogos' in r[8]:
                di = r[8]
                validar_catalogo(pregunta,di,hoja)
            
    
            if totales != 'No':
                
                #if len(totales['total'])==1 and len(totales['subtotal'])>0: #para poner el gris del na a toda la fila
                fn = final_tabla.juntar()
                if not fn:
                        fn =[0]
                poner_gris(letras_validadas,hoja,fn[0])
            
            print('Todo bien con pregunta ', nter)    
            con+=1
            frame.conjuntar_db()
            frame.ajustar_db()
        except:
            print('!!!!no se pudo hacer nada con ',nter)
            traceback.print_exc()
            errores.append(nter)
            con+=1
    return 

def validar_temporal(pregunta, hoja, di):
    diccionario = di['temporales']
    abc = list(string.ascii_uppercase)+['AA','AB','AC','AD']
    
    final = di['final']-1
    if '(aaaa)' in diccionario:
        cor = diccionario['(aaaa)']
        for tupla in cor:
            letra = abc[tupla[1]]
            fila = str(tupla[0]+pregunta+3)
            
            celda = letra+fila
            celdas = emparejar_coordenadas([celda], final)
            dv = DataValidation(type="whole", operator="between",
                                formula1=2000, formula2=2050,
                                allow_blank=True)
            dv.add(celdas[0])
            hoja.add_data_validation(dv)
    
    if '(mm)' in diccionario:
        cor = diccionario['(mm)']
        for tupla in cor:
            letra = abc[tupla[1]]
            fila = str(tupla[0]+pregunta+3)
            
            
            celda = letra+fila
            celdas = emparejar_coordenadas([celda], final)
            dv = DataValidation(type="whole", operator="between",
                                formula1=1, formula2=12,
                                allow_blank=True)
            dv.add(celdas[0])
            hoja.add_data_validation(dv)
    
    if '(dd)' in diccionario:
        cor = diccionario['(dd)']
        for tupla in cor:
            letra = abc[tupla[1]]
            fila = str(tupla[0]+pregunta+3)
            
            
            celda = letra+fila
            celdas = emparejar_coordenadas([celda], final)
            dv = DataValidation(type="whole", operator="between",
                                formula1=1, formula2=31,
                                allow_blank=True)
            dv.add(celdas[0])
            hoja.add_data_validation(dv)
    
    if '(años)' in diccionario:
        cor = diccionario['(años)']
        for tupla in cor:
            letra = abc[tupla[1]]
            fila = str(tupla[0]+pregunta+3)
            
            celda = letra+fila
            celdas = emparejar_coordenadas([celda], final)
            dv = DataValidation(type="whole", operator="between",
                                formula1=0, formula2=99,
                                allow_blank=True)
            dv.add(celdas[0])
            hoja.add_data_validation(dv)
            
    if 'Edad\n(años)' in diccionario:
        cor = diccionario['Edad\n(años)']
        for tupla in cor:
            letra = abc[tupla[1]]
            fila = str(tupla[0]+pregunta+3)
            
            celda = letra+fila
            celdas = emparejar_coordenadas([celda], final)
            dv = DataValidation(type="whole", operator="between",
                                formula1=18, formula2=99,
                                allow_blank=True)
            dv.add(celdas[0])
            hoja.add_data_validation(dv)
    
    return

def gris_desagregados(listaNA,listalistas,hoja):
    """
    

    Parameters
    ----------
    listaNA : lista con las columnas en donde va condicional NA de desagregados para tipo de delitos
    listalistas : list. Lista de listas con tuplas
    hoja : hoja de openpyxl 

    Returns
    -------
    None.

    """
    try:
        para_formula = listaNA[-2]
    except:#excepcion para listas de tipos de delito donde solo hay una columna
        return
    validar = []
    validar.append(listalistas[-1][0])
    for i in listalistas:
        validar += i
    nval=[]
    nval = [valor for valor in validar if valor not in nval]
    nval1 = emparejar_coordenadas(nval,164)
    for celda in nval1:
        hoja.conditional_formatting.add(celda,                              
                                        FormulaRule(formula=[para_formula+'=1'], stopIfTrue=True, fill=gris))
        
    return

def poner_gris(listalistas,hoja,iterar):
    """
    

    Parameters
    ----------
    listalistas : list. Recibe lista de listas con tuplas
    hoja : hoja de excel en openpyxl

    Returns
    -------
    None.
    
    recibe una lista de listas de tuplas para poner condicional gris del NA

    """
    print('entrada poner gris',listalistas)
    validar = []
    validar.append(listalistas[-1][0])
    for i in listalistas:
        validar += i
    nval=[]
    nval = [valor for valor in validar if valor not in nval]
    nval1 = emparejar_coordenadas(nval,iterar)
    for celda in nval1[1:]:
        hoja.conditional_formatting.add(celda,                              
                                        FormulaRule(formula=[nval[0]+'="NA"'], stopIfTrue=True, fill=gris))
        
    return

def soloautosuma(diccionario, part_tab,freal,fila,hoja,autosuma):
    """
    

    Parameters
    ----------
    diccionario : dict. Diccionario con coordenadas (lista de tuplas)
    part_tab : list Lista de lista de tuplas
    freal : int. Numero de fila en dataframe de pandas
    fila : int. Numero de fila en pregunta de dataframe de pandas
    hoja : hoja de excel de openpyxl
    autosuma : fila donde hay que poner la autosuma (es a la fila de la pregunta del dataframe)

    Returns
    -------
    None.

    """
    
    c_dic = [] + diccionario['tuplas']
    if autosuma != 0:
        abc1 = list(string.ascii_uppercase)+['AA','AB','AC','AD']
        if part_tab == 0:
            tuplas = c_dic
        if part_tab != 0:
            tuplas = c_dic
            for i in part_tab:
                for tu in i:
                    a = tu[0] -1 # se hace ajuste por desfaz en las coordenadas 
                    b = tu[1]
                    tuplas.append((a,b))
        if 'sabe' in diccionario:
            tuplas = [tupla for tupla in tuplas if tupla[1] > diccionario['sabe'][-1][1]]
        if 'sabe1' in diccionario:
            tuplas = [tupla for tupla in tuplas if tupla[1] > diccionario['sabe1'][-1][1]]
        autos = freal-fila+autosuma-1
        tuplas.sort()
        col = []
        letcol = []
        for tupla in tuplas:
            if tupla[1] < 30:
                a = tupla[1]
                col.append(a)
            else:
                tuplas.remove(tupla)
        for val in col:
            a = abc1[val]
            letcol.append(a)
        # cords = crearcoordenada1(letcol, tuplas, fila, freal)
        print ('Tablas 3. Autosuma normal: ', letcol, freal, autos+1 )
        c = 0 
        for tupla in tuplas:
            autosumaportupla([letcol[c]], freal, autos, hoja,tupla,fila)
            c += 1
        
    
    return

def emparejar_coordenadas(lista_inicio,numero_fin):
    """
    

    Parameters
    ----------
    lista_inicio : list. Coordenadas de inicio de tabla tipo excel
    numero_fin : int. numero de filas que contiene la tabla

    Returns
    -------
    coords : list. Lista con coordenadas de rangos inicio-fin tipo excel.
    Ejemplo:  'A1:A10'

    """
    
    coords = []
    
    for el in lista_inicio:
        numeroi = [val for val in el if val.isdigit()]
        a = '0'
        for i in numeroi:
            a += i 
        cifra_inicio = int(a)
        listaleras = el.split(str(cifra_inicio))
        letra_columna = listaleras[0]
        cifra_fin = cifra_inicio + numero_fin
        if cifra_fin < cifra_inicio:
            cifra_fin = cifra_inicio
        corrdenada_final = letra_columna + str(cifra_fin) 
        coords.append(el + ':' + corrdenada_final)                    

    return coords

def validarcondicional(diccionario, part_tab,freal,fila,hoja):
    """
    Esta función es para poner las listas desplegables y condicionales

    Parameters
    ----------
    diccionario : dict. Un diccionario que contiene frases
    y coordenadas en listas de tuplas
    part_tab : list. Lista de tuplas de coordenadas sobre partes de tablas
    freal : int. fila real en del dataframe de pandas
    fila : int. fila dentro de la pregunta del dataframe de pandas.
    hoja : hoja de openpyxl.

    Returns
    -------
    None.

    """
    c_tuplas = []+diccionario['tuplas'] #para no editar el diccionario ya que se puede usar en otras funciones
    abc1 = list(string.ascii_uppercase)+['AA','AB','AC','AD']
    if part_tab == 0:
        tuplas = c_tuplas
    if part_tab != 0:
        tuplas = c_tuplas
        for i in part_tab:
            for tu in i:
                a = tu[0] -1 # se hace ajuste por desfaz en las coordenadas 
                b = tu[1]
                tuplas.append((a,b))
    tuplas.sort()
    col = []
    letcol = []
    for tupla in tuplas:
        if tupla[1] < 30:
            a = tupla[1]
            col.append(a)
        else:
            tuplas.remove(tupla)
    for val in col:
        a = abc1[val]
        letcol.append(a)
    cords = crearcoordenada1(letcol, tuplas, fila, freal)
    
    final = diccionario['final1']-1
    ncords = emparejar_coordenadas(cords, final)
    
    if 'sabe' in diccionario:
        # mismo proceso que totales y subtotales, se cancela todo a la derecha hasta que localiza otro si fuera el caso. Hay que hacer iteraciones
        sabe = diccionario['sabe']
        sabetu = []
        cols = []
        for tupla in sabe:
            a = tupla[1]
            b = tuplas[0][1]+1
            sabetu.append((b,a)) #proceso de ajuste para tener misma linea de fila
        for tupla in sabetu:
            a = tupla[1]
            b = abc1[a]
            cols.append(b)
        corval = crearcoordenada1(cols, sabe, fila, freal)
        
        ncor= emparejar_coordenadas(corval,final)
        validar_sabe(hoja,ncor,'1,2,9')
        # función de condicionales
        condic_sabe(hoja,ncor,ncords,1)
    if 'sabe1' in diccionario:
        sabe = diccionario['sabe1']
        sabetu = []
        cols = []
        for tupla in sabe:
            a = tupla[1]
            b = tuplas[0][1]+1
            sabetu.append((b,a)) #proceso de ajuste para tener misma linea de fila
        for tupla in sabetu:
            a = tupla[1]
            b = abc1[a]
            cols.append(b)
        corval = crearcoordenada1(cols, sabe, fila, freal)
        
        ncor= emparejar_coordenadas(corval,final)
        validar_sabe(hoja,ncor,'1,2,3,9')
        # función de condicionales
        condic_sabe(hoja,ncor,ncords,2)
    if 'noaplica' in diccionario:
        sabe = diccionario['noaplica']
        sabetu = []
        cols = []
        for tupla in sabe:
            a = tupla[1]
            b = tuplas[0][1]+1
            sabetu.append((b,a)) #proceso de ajuste para tener misma linea de fila
        for tupla in sabetu:
            a = tupla[1]
            b = abc1[a]
            cols.append(b)
        corval = crearcoordenada1(cols, sabe, fila, freal)
        
        ncor= emparejar_coordenadas(corval,final)
        validar_sabe(hoja,ncor,'1,2,8,9')
        # función de condicionales
        condic_sabe(hoja,ncor,ncords,3)
    if 'NA' in diccionario:
        sabe = diccionario['NA']
        sabetu = []
        cols = []
        for tupla in sabe:
            a = tupla[1]
            b = tuplas[0][1]+1
            sabetu.append((b,a)) #proceso de ajuste para tener misma linea de fila
        for tupla in sabetu:
            a = tupla[1]
            b = abc1[a]
            cols.append(b)
        corval = crearcoordenada1(cols, sabe, fila, freal)
        
        ncor= emparejar_coordenadas(corval,final)
        validar_sabe(hoja,ncor,'X')
        # función de condicionales
        condic_sabe(hoja,ncor,ncords,4)
    
    return

def condic_sabe(hoja, sabe, columnas,codigo):
    "hoja de openpy, sabe: coordenadas de catalogo sinonosabe, columnas: coordenadas de columnas de la tabla. Sabe y columnas deben ser listas de coordenadas"
    no_en_cords = [cor for cor in sabe if cor  not in columnas]
    todas = no_en_cords + columnas
    indice = []
    c = 0

    for cor in todas:
        for i in sabe:
            if i in cor:
                indice.append(c)
        c += 1
    ldl = []
    s = 1
    for i in indice:
        try:
            
            b = indice[s]
            n = todas[i:b]
            if todas[i] == todas[b-1]: #para las que son contiguias que salga toda la fila con el condicional
                n = todas[i:]
            
        except:
            n = todas[i:]
        ldl.append(n)
        s += 1
    
    for lista in ldl:
 
        for ele in lista[1:]:
            alfa = lista[0].split(':')
            lista[0] = alfa[0]
            if codigo == 1:
                hoja.conditional_formatting.add(ele,
                                                FormulaRule(formula=['=OR('+lista[0]+'=2,'+lista[0]+'=9)'], stopIfTrue=True, fill=gris))
            if codigo == 2:
                hoja.conditional_formatting.add(ele,
                                                FormulaRule(formula=['=OR('+lista[0]+'=2,'+lista[0]+'=3,'+lista[0]+'=9)'], stopIfTrue=True, fill=gris))
            if codigo == 3:
                hoja.conditional_formatting.add(ele,
                                                FormulaRule(formula=['=OR('+lista[0]+'=2,'+lista[0]+'=8,'+lista[0]+'=9)'], stopIfTrue=True, fill=gris))
            if codigo == 4:
                hoja.conditional_formatting.add(ele,
                                                FormulaRule(formula=[lista[0]+'="X"'], stopIfTrue=True, fill=gris))
       
    return

def validar_sabe(hoja,celdas,formula):
    "Hoja de openpy y celdas debe ser una lista para iterar con las lertras ya bien definidas"
    dv = DataValidation(type="list", formula1= '"="",{}"'.format(formula), allow_blank=True)

    hoja.add_data_validation(dv)
    cel = celdas[0].split(':')
    a1 = hoja[cel[0]]
    try:
        a1.value = ""
    except:
        print('Habrá un error en los condicionales para la celda', cel[0])
    for celda in celdas:
        dv.add(celda)
    return



def menerrorsub(fila_inicio,fila_suma,hoja,listaletras,fin_pregunta):
    "Función para poner mensaje de error y embudo de errores. los argumentos ya deben tener su ajuste para la escritura"
    w = ['AF','AG','AH','AI','AJ','AK']
    r = ['1','2','4','4','5','6']
    c = 0
    for val in w:
        a = val+str(fila_inicio)
        b = val+str(fila_suma)
        if fila_inicio != fila_suma:
            nomo = a+':'+b
            for lista in listaletras[1:]:
                j = lista[c]+str(fila_inicio)
                k = lista[c]+str(fila_suma)
                nomo += ','+j+':'+k
            que = ['=IF(SUM('+nomo+')>0,'+r[c]+',0)']
            donde = [val+str(fila_suma+1)]
            
        else:
            nomo = a
            for lista in listaletras[1:]:
                j = lista[c]+str(fila_inicio)
                nomo += ','+j
            que = ['=IF(SUM('+nomo+')>0,'+r[c]+',0)']
            donde = [val+str(fila_suma+1)]

        escribirgeneral(donde, que,hoja)
        c+=1
    que1=['=SUM(AF'+str(fila_suma+1)+':'+'AH'+str(fila_suma+1)+')',
          '=SUM(AI'+str(fila_suma+1)+':'+'AK'+str(fila_suma+1)+')']
    donde1 = [ 'AH'+str(fila_suma+2), 'AK'+str(fila_suma+2)]
    escribirgeneral(donde1, que1,hoja)
    que2=['=CAMBIAR('+'AH'+str(fila_suma+2)+',0,"",1,"El total no puede ser cero si se registró un NS o un valor mayor que cero en los desagregados",2,"El total no puede ser NS si registraste algún valor o todos los desagregados son cero",3,"Error en total donde se indica cero y también donde se indica con NS",4,"Cuando el total es NA, la fila se debe dejar en blanco o llenarla con NA según sea la instrucción",5,"Hay error donde el total se indica como cero y también en total con NA",6,"Error donde el total es NS y también donde el total es NA",7,"En los totales que se indica cero, NS y NA ")',
          '=CAMBIAR('+'Ak'+str(fila_suma+2)+',0,"",4,"Hay espacios en blanco que deben ser llenados ",5,"La suma del total no coincide con los desagregados ",6,"Se registraron valores que no son aceptados como respuesta para esta pregunta",9,"Hay espacios en blanco y la suma no coincide con los desagregados",10,"Hay espacio en blanco y valores no admitidos",11,"La suma no coincide con los desagregados y se registran valores no aceptados",15,"Espacios en blanco, suma incorrecta y valores no aceptados")']
    donde2 = ['AH'+str(fila_suma+3), 'AK'+str(fila_suma+3)]  
    escribirgeneral(donde2, que2,hoja)
    try: #escribir debajo de la tabla los mensajes de error
       que3 = ['='+donde2[0],'='+donde2[1]] 
       donde3 = ['C'+str(fin_pregunta+1),'C'+str(fin_pregunta+2)]
       escribir_mensajes_error(donde3, que3,hoja)
    except:
        pass
    return


def variablesnotab(coordenadas,freal,hoja,fin):
    """
    Esta funcion es para poner validaciones en preguntas que no son tablas
    pero que tienen elementos de total y desagregados

    Parameters
    ----------
    coordenadas : list. Lista de tuplas
    freal : int. Fila real del dataframe de pandas
    hoja : hoja de openpyxl.
    fin : int. fin de la pregunta del dataframe para poner autosumas

    Returns
    -------
    None.

    """
    abc1 = list(string.ascii_uppercase)+['AA','AB','AC','AD']
    w = ['AF','AG','AH','AI','AJ','AK','AL']
    nom = ['error_cero','error_NS','error_NA','error_blanco','error_suma',
           'error_valor','hay error fila']
    preg=[]
    agrega = []
    for i in coordenadas:
        letcol= []
        for cor in i:
            a = cor[0]
            b = cor[1] - 4
            fila = a +freal 
            letra =abc1[b]
            co = letra+str(fila)
            letcol.append(co)
            agrega.append(co)
        preg.append(letcol)
    con=0
    agrega = ','.join(agrega)
    frame.agregar_valor_acolumna(agrega,'coordenada')
    for p in preg:
        if len(preg) == 0:
            break
        c = con +freal +5
        c1 = str(c)
        coordss = p
        val_des = coordss[1]+':'+coordss[-1]
        total_tabla = coordss[0]
        Total = 'SUM('+val_des+')'
        fila_res = coordss[0]+':'+coordss[-1]
        NS = 'COUNTIF('+fila_res+',"NS")'
        coincide = 'IF(OR('+Total+'='+total_tabla+','+total_tabla+'="",AND('+total_tabla+'>0,'+NS+'='+str(len(coordss)-1)+')),0,1)'
        blancos = parablancos(coordss)
        aha = str(len(coordss))
        hay='AF'+c1+':AK'+c1
        texto = 'COUNTIF('+fila_res+',"=*")'
        NA = 'COUNTIF('+fila_res+',"NA")'
        formulas = formula_aritmetica(
            total_tabla,
            Total,
            NS,
            NA,
            blancos,
            aha,
            coincide,
            hay,
            texto
            )
        donde = []
        for i in w:
            y = i+c1
            donde.append(y)
        que = []
        for i in nom:
            a = formulas[i]
            que.append(a)
        escribirgeneral(donde, que, hoja)
        for cor in coordss:
            hoja.conditional_formatting.add(cor,
                                            FormulaRule(formula=[donde[-1]+'=1'], stopIfTrue=True, fill=redFill))

        con+=1
        
    en =[]
    for i in w:
        y = i+str(freal+4)
        en.append(y)
    if len(preg)>0:
        escribirgeneral(en, nom,hoja)
        menerror_no_tab(freal+5,freal+5+con,hoja,fin+freal-2)
    return 

def borrar_tuplas(diccionario, tuplas):
    """
    

    Parameters
    ----------
    diccionario : dict. diccionario arrojado en r[8] de analizarcor
    tuplas : list. Lista de tuplas de columnas

    Returns
    -------
    tuplas : list. nueva lista de tuplas de columnas sin 
    las tuplas del sabenosabe etc

    """
    
    if len(diccionario) > 1:
        tuplas_borrar = []
        if 'sabe' in diccionario:
            tuplas_borrar += diccionario['sabe']
        if 'sabe1' in diccionario:
            tuplas_borrar += diccionario['sabe1']
        # if 'noaplica' in diccionario:       esas dos suelen ir antes de los totales a validar
        #     tuplas_borrar += diccionario['noaplica']
        # if 'NA' in diccionario:
        #     tuplas_borrar += diccionario['NA']
        for borrar in tuplas_borrar:
            for objetivo in tuplas:
                a = range(objetivo[0]-2,objetivo[0]+2)
                if borrar[0] in a:
                    if borrar[1] == objetivo[1]:
                        tuplas.remove(objetivo)
    else:
        tuplas = tuplas
    return tuplas

def subtuplas(totales, tuplas, part_tab, dic):
    """
    

    Parameters
    ----------
    totales : dict. Diccionario con listas de tuplas donde se encontró 
    total o subtotal en la tabla
    tuplas : list. List de tuplas para comparar con totales y 
    subtotales (agregar, quitar, o clasificar elementos)
    part_tab : list. Lista de listas de tuplas para tablas que van en partes
    dic : dict. Diccionario con todas tuplas y coordenadas de frases 
    para catálogo, o donde van listas desplegables
    Returns
    -------
    Regresa siempre una lista de listas con las tuplas por cada total o 
    subtotal dentro de la tabla, con la finalidad de poder iterarlas
    y hacer la validación para todos los elementos.

    """
    
    print('entrada subtuplas: ', totales, tuplas, part_tab)
    total = totales['total']
    sub = totales['subtotal']
    
    if len(total) > 0: #este if es po si el total no está dentro de las tuplas, para incluirlo. 
        at = [tupla[1] for tupla in tuplas]
        if total [0][1] not in at:
            tuplas = [(tuplas[0][0],total[0][1])]+tuplas
    tuplas = borrar_tuplas(dic, tuplas)
    
    if len(sub) == 0 and len(total) == 1 and part_tab != 0: #para tablas con un solo total y part_tab
        nuevas_tuplas = [] + tuplas
        for lista in part_tab:
            nuevas_tuplas += lista
        nuevas_tuplas.append('a')
        return [nuevas_tuplas]
    
    if len(total) > 1 and len(sub) > 0 and part_tab == 0: #Aqui para totales > 1 y subtotales en medio. Solo para un nivel de tabla
        
        todas_tuplas = []+tuplas
        listalistas=[]
        listalistas.append(tuplas)
        niveles_totales = [tupla[0] for tupla in total]
        nt_unicos = list(set(niveles_totales))

        # if part_tab != 0:
        #     for lista in part_tab:
        #         todas_tuplas += lista
        #         listalistas.append(lista)
        if len(nt_unicos) == 1: #solo hay un nivel en unicos de totales
            sub_residual = []
            tot_residual = [] #lista de listas y contiene todas tuplas por total
            col_tot = [tupla[1] for tupla in total]
            for lista in listalistas:
                c = 1
                
                for val in col_tot:
                    t1 = []
                    for tupla in lista:
                        try:
                            if tupla[1] in range(val,col_tot[c]):
                                t1.append(tupla)
                        except:
                            if tupla [1] >= val:
                                t1.append(tupla)
                    
                    tot_residual.append(t1)
                

                    t2 = []
                    for tupla in sub:
                        try:
                            if tupla[1] in range(val,col_tot[c]):
                                t2.append(tupla)
                        except:
                            if tupla [1] >= val:
                                t2.append(tupla)
                    c += 1
                    sub_residual.append(t2)
            
                
            
            c = 0
            coltu =[]
            for lista in tot_residual: # REcursividad!!!! Cuidado!!!
                nd = {'total':[total[c]],'subtotal':sub_residual[c]}
                
                aderir = subtuplas(nd,lista,0,dic)
                for i in aderir:
                    coltu.append(i)
                c += 1
            
            return coltu
  
    if len(sub) == 0 and len(total) > 1 and part_tab == 0:
        resta = total[1][0] - total[0][0]
        if resta > 3:
            return tuplas
        else:
            col = []
            coltu = []
            for i in total:
                a = i[1]
                col.append(a)
            c = 0
            for i in col:          
                cm = []  
                try:
                    z = range(i,col[c+1])
                    for tupla in tuplas:
                        q = tupla[1]
                        if q in z:
                            cm.append(tupla)
                except:
                    for tupla in tuplas:
                        q = tupla[1]
                        if q >= i:
                            cm.append(tupla)   
                c+=1
                coltu.append(cm)
            return coltu
    if len(total) == 1 and len(sub) > 1 and part_tab != 0:
        #comprobacion adicional de tublas de primer nivel --- eliminar segundos y/terceros niveles de preguntas por partes
        # com =[]
        # for tupla in tuplas:
        #     com.append(tupla[0])
        # nu = list(set(com))
        # print('nu',nu)
        # if len(nu)>1:
        #     nu_men = min(nu) 
        #     print('nu_men',nu_men)
        #     for tupla in tuplas:
        #         if tupla[0] > nu_men:
        #             tuplas.remove(tupla)
        #filas unicas por alguna razón sale así y hay que borrar
        
        todas_tuplas = []+tuplas
        listalistas=[]
        listalistas.append(tuplas)
        print('lis',listalistas)
        for lista in part_tab:
            todas_tuplas += lista
            listalistas.append(lista)
        subtotales = []
        niveles = []
        for tupla in todas_tuplas:
            a = tupla[0]
            niveles.append(a)
        unicos_niv = list(set(niveles))
        unicos_niv.sort()
        coltu = []
        for niv in unicos_niv:
            l = []
            for tupl in sub:
                a = tupl[0]
                if a in range(niv-3,niv+3):
                    l.append(tupl)
            subtotales.append(l)
        p = 0
        ncol = []
        ncol1 = []
        for sub in subtotales:
            col = []
            for i in sub:
                a = i[1]
                col.append(a)
                ncol.append(a)
            c = 0
            ncol1.append(col)
            for i in col:          
                cm = []  
                try:
                    z = range(i,col[c+1])
                    for tupla in listalistas[p]:
                        q = tupla[1]
                        if q in z:
                            cm.append(tupla)
                except:
                    for tupla in listalistas[p]:
                        q = tupla[1]
                        if q >= i:
                            cm.append(tupla)   
                c+=1
                coltu.append(cm)
            p += 1
        to = total[0][1]
        print('esto pasa',to,ncol[0])
        extras = [tupla for tupla in tuplas if tupla[1] in range(to,ncol[0])]
        if len(extras) > 0:
            len_ex = len(extras) - 1 #menos uno para ignorarl el total
            os = []
            conta = 1
            print('aaaaaaaa', coltu, extras)
            for tupla in extras[1:]:
                ti = [tupla]
                for lista in coltu:
                    len_lista = len(lista) - 1 #para solo saber de sus desagregados
                    if len_lista != len_ex:
                        veces = len_lista / len_ex
                        salto = len_lista / veces
                        salto = int(salto)
                        for tupla1 in lista[conta::salto]:
                            ti.append(tupla1)
                    else:
                        ti.append(lista[conta])
                conta += 1
                ti.append('a') #para subttales que no se llama así, el A es para identificarles, pero se borra antes de ingresar a la funcion de validacion
                os.append(ti)
            for lista in os:
                coltu.append(lista)
                
        tt = [(tuplas[0][0],total[0][1])]
        c = 0
        for tups in listalistas:
            for tup in tups: #para el total con subtotales y otros que queden fuera
                w = tup[1]
                for val in ncol1[c]:
                    if w == val:
                        tt.append(tup)
            c += 1
        tt.append('a')
        coltu.append(tt)
        # print('qqqqqqqqq', coltu,ncol1)
        return coltu
    else:
        if len(sub) == 0: #si es tabla de varios totales
            col = []
            coltu = []
            for i in total:
                a = i[1]
                col.append(a)
            c = 0
            for i in col:          
                cm = []  
                try:
                    z = range(i,col[c+1])
                    for tupla in tuplas:
                        q = tupla[1]
                        if q in z:
                            cm.append(tupla)
                except:
                    for tupla in tuplas:
                        q = tupla[1]
                        if q >= i:
                            cm.append(tupla)   
                c+=1
                coltu.append(cm)

        if len(sub) > 0: #si es tabla de varios subtotales
            col = []
            coltu = []
            for i in sub:
                a = i[1]
                col.append(a)
            c = 0
            col.sort()
            for i in col:          
                cm = []  
                try:
                    z = range(i,col[c+1])
                    for tupla in tuplas:
                        q = tupla[1]
                        if q in z:
                            cm.append(tupla)
                except:
                    for tupla in tuplas:
                        q = tupla[1]
                        if q >= i:
                            cm.append(tupla)   
                c+=1
                coltu.append(cm)
                
            to = total[0][1]
            extras = [tupla for tupla in tuplas if tupla[1] in range(to,col[0])]
            if len(extras) > 0:
                len_ex = len(extras) - 1 #menos uno para ignorarl el total
                os = []
                conta = 1
                # print('aaaaaaaa', coltu, extras)
                for tupla in extras[1:]:
                    ti = [tupla]
                    for lista in coltu:
                        len_lista = len(lista) - 1 #para solo saber de sus desagregados
                        if len_lista != len_ex:
                            veces = len_lista / len_ex
                            salto = len_lista / veces
                            salto = int(salto)
                            for tupla1 in lista[conta::salto]:
                                ti.append(tupla1)
                        else:
                            ti.append(lista[conta])
                    conta += 1
                    ti.append('a') #para subttales que no se llama así, el A es para identificarles, pero se borra antes de ingresar a la funcion de validacion
                    os.append(ti)
                for lista in os:
                    coltu.append(lista)
  
            tt = [(tuplas[0][0],total[0][1])]
            for tupla in tuplas: #para el total con subtotales y otros que queden fuera
                w = tupla[1]
                for val in col:
                    if w == val:
                        tt.append(tupla)
            tt.append('a')
            coltu.append(tt)
        return coltu
    

def clasi(tuplas):
    "para clasificar y procesar las que no son tablas"
    mayores=[]
    for tupla in tuplas:
        r = tupla[1]
        if r > 3:
            mayores.append(tupla)
    jer = jerarquia(mayores) #len es el numero de niveles de la pregunta
    primero = []
    segundo = []
    tercero = []
    cuarto = []
    iterar = [primero,segundo,tercero,cuarto]
    if len(jer) == 1:
        pass #construir despues con blancos y errores de valor
    if len(jer) > 1:
        r = 0
        for nivel in jer:
            
            for i in mayores:
                if jer[r] == i[1]:
                   iterar[r].append(i) 
            r+=1
    test = niveles(iterar)
    try:
        da = niveles(iterar[1:])
        for i in da:
            test.append(i)
    except:
        pass
    try:
        de = niveles(iterar[2:])
        for i in de:
            test.append(i)
    except:
        pass
    nuy = []
    for i in test:
        if len(i)>1:
            nuy.append(i)
    return nuy

def niveles(lista_de_listas):
    "Definir niveles en preguntas que no son tablas"
    res = []
    if len(lista_de_listas[0])>1:
        c=0 
        for i in range(0,len(lista_de_listas[0])-1):
            p = lista_de_listas[0]
            x = p[c]
            xa = x[0]
            ad = lista_de_listas[0]
            ad1 = ad[c+1]
            x1 = ad1[0]
            esp = x1-1
            lis = [x]
            for vl in lista_de_listas[1]:
                x2 = vl[0]
                if x2 < esp and x2 > xa:
                    lis.append(vl)
            c+=1
            res.append(lis)
        try:
            p = lista_de_listas[0]
            q = p[-1]
            x = q[0]
            lid = [q]
            for v in lista_de_listas[1]:
                x2 = v[0]
                if x2 > x:
                    lid.append(v)
            if len(lid)>1:
                res.append(lid)
        except:
            pass
    if len(lista_de_listas[0])==1:
        
        p = lista_de_listas[0]
        q = p[0]
        x = q[0]
        lid = [q]
        for v in lista_de_listas[1]:
            x2 = v[0]
            if x2 > x:
                lid.append(v)
        if len(lid)>1:
            res.append(lid)
    return res
        

def jerarquia(lista):
    "definir jerarquía para preguntas que no son tablas"
    uni = []
    for tupla in lista:
        a = tupla[1]
        uni.append(a)
    nu = list(set(uni))
    nu.sort()
    return nu

def listalistas(lista):
    "Regresa una lista de listas a partir de separar valores de la lista dada como argumento"
    listas=[]
    indices = []
    c = 0
    for i in lista:
        try:
            a = lista[c+1]-lista[c]
            if a > 1:
                indices.append(c+1)
        except:
            pass
        c+=1
    d = 0
    for i in indices:
        if d == 0:
            li = lista[d:i]
            listas.append(li)
        else:
            lim = indices[d-1]
            li =lista[lim:i]
            listas.append(li)
            if i == indices[-1]:
                li = lista[i:]
                listas.append(li)
        d+=1
        
    return listas



def condicional(letcol,freal,hoja,fila, tuplas):
    "Columna y letras deben ser listas, freal es un int, freal no debe estar ajustado. Regresa lista de string con rango de desagregados en fila"
    r = []
    c = 0
    for tupla in tuplas:
        ftupla = tupla[0]
        if ftupla-fila > 0:
            ftupla += 1
        inicio = freal - fila + ftupla
        
        s = letcol[c] + str(inicio+2)
        r.append(s)
        c += 1
    return r

def vardelitos(fila,tuplas,freal,coldelito,codelitos,fpregunta,autosu,hoja,part_tab,letras,men):
    "Función para poner validación en tablas de tipos de delito"
    print('entrada vardelitos: ', fila,tuplas,freal,coldelito,codelitos,fpregunta,autosu )
    codelitos1=listalistas(codelitos)
    abc1 = list(string.ascii_uppercase)+['AA','AB','AC','AD']
    if len(tuplas) == 1:
        print('seguro es un subtotal unico')
        r = [abc1[tuplas[0][1]]+str(freal+2)]
        return r
    ntuplas = [] #hacer ajuste de las tuplas de otras filas
    copia_tuplas = []
    for tupla in tuplas:
        if tupla[0] > fila:
            a = (tupla[0]-2,tupla[1])
            copia_tuplas.append(a)
        if tupla[0] < fila or tupla[0] == fila :
            copia_tuplas.append((fila-2,tupla[1]))
        a = (tupla[0]-1,tupla[1])
        ntuplas.append(a)
    tuplas = ntuplas 
  
    columnas = []
    letcol= []
    for tupla in tuplas:
        a = tupla[1]
        columnas.append(a)
    for i in columnas:
        s = abc1[i]
        letcol.append(s)
    formulas = ['coincide_con_desagregados?',
                'hay NA',
                'mayor al 25%']
    x1 = variables(fila-2,copia_tuplas,freal,autosu-2,hoja,part_tab,letras,men)
    #166 como autosuma que se retirará de la funcion variables para que no cause error
    ax1 = x1[0]
    x1 = x1[1]
    c = 0
    for i in letcol:
        autosumadelito([i],fpregunta,hoja, tuplas[c])
        c += 1
       
    totls = totls1(codelitos1)
    delito_total=deltot(codelitos1,letcol[0],fpregunta,tuplas[0],fila)
    sus_desagregados=desagregadoss(codelitos1,letcol[0],fpregunta,tuplas[0],fila)
    az = []
    u = x1[-1]
    lc = u[:2]
    letrasesc= nl()
    we= letrasesc.index(lc)
    az.append(letrasesc[we+1])
    az.append(letrasesc[we+2])
    az.append(letrasesc[we+3])
    todo= []
    condi = [] #columnas ara condicionales, son 3 y ya va con fila real (Son las de la validación de la derecha de las tablas)
    for i in az:
        tr = i+str(freal+1)
        todo.append(tr)
        tr1 = i+str(freal+2)
        condi.append(tr1)
    
    escribirgeneral(todo,formulas, hoja)
    condi1 = [] #lista de letras que tienen los numeros de la fila
    condi2 = []
    for l in letcol:
        condi1.append(lw(l,freal+(tupla[0]-fila)+2,2))
        condi2.append(lw(l,freal+(tupla[0]-fila)+166,2))
    x_valida.extraer(condi)
    condi3 = []
    c = 0
    for letra in condi1:
        condi3.append(letra+':'+condi2[c])
        c += 1
    
    condi1 = condi3
    hoja.conditional_formatting.add(condi1[0],
                                    FormulaRule(formula=[condi[0]+'=1'], stopIfTrue=True, fill=redFill))
    hoja.conditional_formatting.add(condi1[0],
                                    FormulaRule(formula=[condi[2]+'=1'], stopIfTrue=True, fill=azul))
    for colu in condi1:
        
        hoja.conditional_formatting.add(colu,
                                        FormulaRule(formula=[condi[1]+'=1'], stopIfTrue=True, fill=gris))
        # if colu != condi1[0]:
        #     hoja.conditional_formatting.add(colu,
        #                                     FormulaRule(formula=[condi1[0]+'="NA"'], stopIfTrue=True, fill=gris))

    coin=az[0]
    ro=0
    for i in delito_total:
        totl=i
        low=sus_desagregados[ro]
        desag = low[0]+':'+low[-1]
        desa='SUM('+desag+')'
        #=SI(O(AS377=AT377,Y(AS377="NA",AT377=0),Y(AS377="NS",AT377=0)),SI(Y(AS377="NA",AT377=0),"NA",0),1)
        escri= ['=IF(OR('+totl+'=SUM('+desag+'),AND('+totl+'="NA",'+desa+'=0),AND('+totl+'="NS",'+desa+'=0),AND('+totl+'>0,SUM('+desag+')=0,COUNTIF('+desag+',"NS")>0)),IF(AND('+totl+'="NA",'+desa+'=0),"NA",0),1)']
        qw=totls[ro]+fpregunta #es fila de pregunta no real
        ax=[coin+str(qw)]
        escribirgeneral(ax, escri, hoja)
        hoja.conditional_formatting.add(desag,
                                        FormulaRule(formula=['$'+coin+'$'+str(qw)+'=1'], stopIfTrue=True, fill=redFill))
        ro+=1
    #para segunda columna, de NA
    beta=az[1]
    listasdesagregados=desagregadoss(codelitos1, beta, fpregunta,tuplas[0],tuplas[0][0]) #doble tupla para que d´cero, igual en la de abajjo ya que es para escribir en las columnas correspondientes
    gama=0
    for lista in listasdesagregados:
        b = delito_total[gama]
        for fila1 in lista:
            donde=[fila1]
            escr = ['=IF('+b+'="NA",1,0)']
            escribirgeneral(donde, escr, hoja)
        gama+=1
    
    delta=az[2]
    listasdesagregados1=desagregadoss(codelitos1, delta, fpregunta, tuplas[0],tuplas[0][0])
    ya=0
    for lista in  listasdesagregados1:
        b = delito_total[ya] 
        ar = sus_desagregados[ya]
        ar1 = ar[-2]
        pri = [lista[-2]]
        escr = ['=IF('+ar1+'/'+b+'>0.25,1,0)']
        escribirgeneral(pri, escr, hoja)
        ya+=1
    # escr = ['=IF(SUM('+x1[1]+':'+x1[-3]+')>0,1,0)'] #estas lineas son para formula de hay
    # donde=[x1[-1]]
    # escribirgeneral(donde, escr,hoja)    
    
    return ax1

def totls1(lista_de_listas):
    "Extrae los primeros valores de las listas que hay dentro de la lista de lista"
    lista=[]
    for i in lista_de_listas:
        r= i[0]
        f= r-1
        lista.append(f)
    return lista

def listadecor(letra, lista_de_listas,freal):
    "regresa una lista con las coordenadas tipo excel"
    resul=[]
    for lista in lista_de_listas:
        to=[]
        for valor in lista:
            w =  freal+2+valor
            a = letra+str(w)
            to.append(a)
        resul.append(to)
    return resul


def desagregadoss(lista_de_listas, letra,freal, tupla, fila):
    "Genera lista con coordenadas tipo excel"
    ta = tupla[0] - fila
    if ta != 0:
        ta += 2
    re=[]
    for i in lista_de_listas:
        sa=[]
        for valor in i: 
            a= valor
            b=a+freal + ta
            c=letra+str(b)
            sa.append(c)
        re.append(sa)
    return re

def deltot(lista_de_listas, letra,freal,tupla,fila):
    "Genera coordenadas tipo excel"
    ta = tupla[0] - fila
    if ta != 0:
        ta += 2
    re=[]
    for i in lista_de_listas:
        a= i[0]
        b=a-1+freal + ta
        c=letra+str(b)
        re.append(c)
    return re
    


def variablesl1(fila,tuplas,freal,autosuma,hoja,letras):
    "Función para validar tablas de una sola columna de respuesta"
    print('valores de variablel1-:', fila,tuplas,freal,autosuma)
    abc1 = list(string.ascii_uppercase)+['AA','AB','AC','AD']
    columnas = []
    letcol= []
    for tupla in tuplas:
        if fila in tupla:
            a = tupla[1]
            columnas.append(a)
    for i in columnas:
        s = abc1[i]
        letcol.append(s)
    coordss = crearcoordenada(letcol, freal)
    
    fila_res = coordss[0]
    NS = 'COUNTIF('+fila_res+',"NS")'
    texto = 'COUNTIF('+fila_res+',"=*")'
    NA = 'COUNTIF('+fila_res+',"NA")'
    formulas = {'errorval':'=IF(OR('+texto+'='+NA+','+texto+'='+NS+'),0,1)',
                'blanco':'=IF('+fila_res+'="",1,0)'}
    
    escribir(freal+2,formulas,hoja,letras)
    autosumanormal(letcol, freal, autosuma, hoja)
    menerrorc1(freal+2, autosuma+2,hoja)
    return

def menerrorc1(fila_inicio,fila_suma,hoja):
    "Función para poner mensajes de error en tablas de columna única de respuesta"
    print('menerrorc1: ',fila_inicio,fila_suma)
    "los argumentos ya deben tener su ajuste para la escritura"
    w = ['AF','AG']
    for val in w:
        a = val+str(fila_inicio)
        b = val+str(fila_suma)
        que = ['=SUM('+a+':'+b+')']
        donde = [val+str(fila_suma+1)]
        escribirgeneral(donde, que,hoja)
    que1=['=IF(AF'+str(fila_suma+1)+'>0,"El valor introducido no es aceptado en este tipo de pregunta","")',
          '=IF(AND(AG'+str(fila_suma+1)+'>0,AG'+str(fila_suma+1)+'<'+str(fila_suma-fila_inicio)+'),"Hay espacios en blanco","")']
    donde1 = [ 'AF'+str(fila_suma+2), 'AG'+str(fila_suma+2)]
    escribirgeneral(donde1, que1,hoja)
    return

def autosumanormal(columnas, inicio,fin,hoja):
    "Función para poner autosumas"
    ini = crearcoordenada(columnas, inicio)
    fina = crearcoordenada(columnas, fin-1)
    donde = crearcoordenada(columnas, fin)
    que = []
    c = 0
    for i in ini:
        b = ini[c]+':'+fina[c]
        a = '=IF(AND(SUM('+b+')=0,COUNTIF('+b+',"NS")>0),"NS",IF(AND(SUM('+b+')=0, COUNTIF('+b+',"NA")>0),"NA",SUM('+b+')))'
        que.append(a)
        c+=1
    dn = ','.join(donde)
    frame.agregar_valor_acolumna(dn,'coordenada')
    try:
        escribirgeneral(donde, que, hoja)
    except:
        pass
    return

def autosumaportupla(columnas, freal, autosuma, hoja, tupla, fila):
    "Función para poner autosuma en preguntas que no son tuplas continuas. columnas es la letra de la columna"
    ftupla = tupla[0]
    if ftupla-fila > 0:
        ftupla += 1
    inicio = freal - fila + ftupla
    fin = inicio + (autosuma - freal)
    ini = crearcoordenada(columnas, inicio)
    fina = crearcoordenada(columnas, fin-1)
    donde = crearcoordenada(columnas, fin)
    que = []
    c = 0
    for i in ini:
        b = ini[c]+':'+fina[c]
        a = '=IF(AND(SUM('+b+')=0,COUNTIF('+b+',"NS")>0),"NS",IF(AND(SUM('+b+')=0, COUNTIF('+b+',"NA")>0),"NA",SUM('+b+')))'
        que.append(a)
        c+=1
    dn = ','.join(donde)
    frame.agregar_valor_acolumna(dn,'coordenada')
    try:
        escribirgeneral(donde, que, hoja)
    except:
        pass
    return




def autosumadelito(columnas, pregunta,hoja,tupla):
    "Función para poner autosumas en tablas de tipo de delito"
    ftupla = tupla[0]
    inicio = pregunta + ftupla
    fin = inicio + 165
    donde = crearcoordenada(columnas, fin)
    L=columnas[0]
    inicio += 2
    k= str(inicio)
    que = ['=IF(AND(SUM(COUNTIF('+L+k+':'+lw(L,inicio,10)+',"NS"),COUNTIF('+lw(L,inicio,17)+':'+lw(L,inicio,21)+',"NS"),'+ #
                      'COUNTIF('+lw(L,inicio,27)+':'+lw(L,inicio,30)+',"NS"),COUNTIF('+lw(L,inicio,49)+',"NS"),'+ #
                      'COUNTIF('+lw(L,inicio,55)+':'+lw(L,inicio,57)+',"NS"),COUNTIF('+lw(L,inicio,61)+':'+lw(L,inicio,67)+',"NS"),'+ #
                      'COUNTIF('+lw(L,inicio,75)+',"NS"),COUNTIF('+lw(L,inicio,81)+':'+lw(L,inicio,85)+',"NS"),'+ #
                      'COUNTIF('+lw(L,inicio,92)+',"NS"),COUNTIF('+lw(L,inicio,101)+':'+lw(L,inicio,104)+',"NS"),'+ #
                      'COUNTIF('+lw(L,inicio,111)+':'+lw(L,inicio,114)+',"NS"),COUNTIF('+lw(L,inicio,124)+',"NS"),'+ #
                      'COUNTIF('+lw(L,inicio,128)+':'+lw(L,inicio,135)+',"NS"),COUNTIF('+lw(L,inicio,143)+':'+lw(L,inicio,144)+',"NS"),'+ #
                      'COUNTIF('+lw(L,inicio,148)+':'+lw(L,inicio,156)+',"NS"),COUNTIF('+lw(L,inicio,162)+':'+lw(L,inicio,164)+',"NS"))>0,'+
                 'SUM('+L+k+':'+lw(L,inicio,10)+','+lw(L,inicio,17)+':'+lw(L,inicio,21)+','+lw(L,inicio,27)+':'+lw(L,inicio,30)+','+lw(L,inicio,49)+','+lw(L,inicio,55)+':'+lw(L,inicio,57)+','+lw(L,inicio,61)+':'+lw(L,inicio,67)+','+
                      lw(L,inicio,75)+','+lw(L,inicio,81)+':'+lw(L,inicio,85)+','+lw(L,inicio,92)+','+lw(L,inicio,101)+':'+lw(L,inicio,104)+','+lw(L,inicio,111)+':'+lw(L,inicio,114)+','+lw(L,inicio,124)+','+lw(L,inicio,128)+':'+lw(L,inicio,135)+','+
                      lw(L,inicio,143)+':'+lw(L,inicio,144)+','+lw(L,inicio,148)+':'+lw(L,inicio,156)+','+lw(L,inicio,162)+':'+lw(L,inicio,164)+')=0),"NS",'+
               'SUM('+L+k+':'+lw(L,inicio,10)+','+lw(L,inicio,17)+':'+lw(L,inicio,21)+','+lw(L,inicio,27)+':'+lw(L,inicio,30)+','+lw(L,inicio,49)+','+lw(L,inicio,55)+':'+lw(L,inicio,57)+','+lw(L,inicio,61)+':'+lw(L,inicio,67)+','+
                      lw(L,inicio,75)+','+lw(L,inicio,81)+':'+lw(L,inicio,85)+','+lw(L,inicio,92)+','+lw(L,inicio,101)+':'+lw(L,inicio,104)+','+lw(L,inicio,111)+':'+lw(L,inicio,114)+','+lw(L,inicio,124)+','+lw(L,inicio,128)+':'+lw(L,inicio,135)+','+
                      lw(L,inicio,143)+':'+lw(L,inicio,144)+','+lw(L,inicio,148)+':'+lw(L,inicio,156)+','+lw(L,inicio,162)+':'+lw(L,inicio,164)+'))']

    dn = ','.join(donde)
    frame.agregar_valor_acolumna(dn,'coordenada')
    escribirgeneral(donde, que,hoja)
    
    return

def lw(l,i,n):
    "Generar coordenada tipo excel, ejemplo: A1"
    a = l+str(n+i)
    return a


def masdeunatablauni1(tuplas):
    val = []
    for tupla in tuplas:
        a = tupla[0]
        val.append(a)
    val1 = []
    for i in val:
        b = val.count(i)
        if b > 2:
            val1.append(i)
    v = list(set(val1))
    return v


def masdeunatablauni(tuplas):
    val = []
    for tupla in tuplas:
        a = tupla[0]
        val.append(a)
    v = list(set(val))
    return v

def formula_aritmetica(
        total_tabla,
        Total,
        NS,
        NA,
        blancos,
        aha,
        coincide,
        hay,
        texto):
    formula = {
        'error_cero':'=IF(AND('+total_tabla+'=0,OR('+Total+'>0,'+NS+'>0)),1,0)',
        'error_NS':'=IF(OR(AND('+total_tabla+'="NS",'+Total+'>0),AND('+total_tabla+'="NS",'+NS+'<2)),1,0)',
        'error_NA':'=IF(AND('+total_tabla+'="NA",OR('+Total+'>0,'+NS+'>0,AND('+NA+'>1,'+NA+'<'+aha+'))),1,0)',
        'error_blanco':'=IF(AND('+blancos+'>0,'+blancos+'<'+aha+','+total_tabla+'<>"NA"),1,0)',
        'error_suma':'=IF(AND('+coincide+'=1,'+total_tabla+'<>"NS",'+total_tabla+'<>"NA"),1,0)',
        'error_valor':'=IF('+texto+'<>SUM('+NS+','+NA+'),1,0)',
        'hay error fila':'=IF(SUM('+hay+')>0,1,0)'
        }
    return formula

def variables(fila,tuplas,freal,autosuma,hoja,part_tab,letras,men):
    "part tab es si la tabla está en partes, men es si o no, y es para el mensaje de errores debido al proceso extrerno si hay totales o subtotales en las tablas"
    print('argumentos de entrada funcion variable',fila,tuplas,freal,autosuma )
    abc1 = list(string.ascii_uppercase)+['AA','AB','AC','AD']
    if len(tuplas) == 1:
        print('seguro es un subtotal unico')
        r = [[abc1[tuplas[0][1]]+str(freal+2)]]
        return r
    columnas = []
    letcol= []
    tuplas.sort()
    for tupla in tuplas:
        a = tupla[1]
        columnas.append(a)
    for i in columnas:
        s = abc1[i]
        letcol.append(s)
    alt1 = []
    fcon1 = []
    iterar = autosuma - freal - 1
    if iterar > 0:
        iterar += 1
    else:
        iterar = 1
    final_tabla.extraer([iterar])
    dfreal = freal + 0
    chek = []
    for i in range(0,iterar):
        freal = dfreal + i
        coordss = crearcoordenada1(letcol,tuplas,fila,freal)
        ccoordss = crearcoordenada1(letcol,tuplas,fila,freal)
        chek.append(coordss)
        if part_tab == 0:
            val_des = coordss[1]+':'+coordss[-1]
            total_tabla = coordss[0]
            Total = 'SUM('+val_des+')'
            fila_res = coordss[0]+':'+coordss[-1]
            NS = 'COUNTIF('+fila_res+',"NS")'
            aha = str(len(coordss))
            coincide = 'IF(OR('+Total+'='+total_tabla+','+total_tabla+'="",AND('+total_tabla+'>0,'+NS+'='+str(len(coordss)-1)+')),0,1)'
            blancos = parablancos(coordss)
            texto = 'COUNTIF('+fila_res+',"=*")'
            NA = 'COUNTIF('+fila_res+',"NA")'
            hay = 'IF(AND('+total_tabla+'=0,'+Total+'>0,'+NS+'=0),1,0),IF(OR(AND('+total_tabla+'="NS",'+Total+'>0),AND('+total_tabla+'="NS",'+NS+'<2)),1,0),=IF(AND('+total_tabla+'="NA",OR('+Total+'>0,'+NS+'>0)),1,0),IF(AND('+blancos+'>0,'+blancos+'<'+aha+','+total_tabla+'<>"NA"),1,0),IF(AND('+coincide+'=1,'+total_tabla+'<>"NS",'+total_tabla+'<>"NA"),1,0),IF('+texto+'<>SUM('+NS+','+NA+'),1,0)'
            formulas = formula_aritmetica(
                total_tabla,
                Total,
                NS,
                NA,
                blancos,
                aha,
                coincide,
                hay,
                texto
                )
        else:
            print('tuplas de tabla en partes: ',part_tab)
            aderir = []
            ncorr = ccoordss
            for lista in part_tab:
                letcol1 = []
                fini = freal - fila
                columnas1 = []
                for tupla in lista:
                    aq = tupla[1]
                    columnas1.append(aq)
                for i in columnas1:
                    s1 = abc1[i]
                    letcol1.append(s1)
                cor = crearcoordenada(letcol1, fini + lista[0][0])
                ncorr += cor
                tor = cor[0]+':'+cor[-1]
                aderir.append(tor)
            val_des = coordss[1]+':'+coordss[-1]
            total_tabla = coordss[0]
            
            fila_res = coordss[0]+':'+coordss[-1]
            
            blancos = parablancos(ncorr)
            aha = str(len(ncorr))
            aha1 = str(len(ncorr)-1)
            NS = 'COUNTIF('+fila_res+',"NS")'
            texto = 'COUNTIF('+fila_res+',"=*")'
            NA = 'COUNTIF('+fila_res+',"NA")'
            for i in aderir:
                val_des += ','+i
                NS += '+COUNTIF('+i+',"NS")'
                texto += '+COUNTIF('+i+',"=*")'
                NA += '+COUNTIF('+i+',"NA")'
            Total = 'SUM('+val_des+')'
            coincide = 'IF(OR('+Total+'='+total_tabla+','+total_tabla+'="",AND('+total_tabla+'>0,'+NS+'='+aha1+')),0,1)'
            hay = 'IF(AND('+total_tabla+'=0,'+Total+'>0,'+NS+'=0),1,0),IF(OR(AND('+total_tabla+'="NS",'+Total+'>0),AND('+total_tabla+'="NS",'+NS+'<2)),1,0),=IF(AND('+total_tabla+'="NA",OR('+Total+'>0,'+NS+'>0)),1,0),IF(AND('+blancos+'>0,'+blancos+'<'+aha+','+total_tabla+'<>"NA"),1,0),IF(AND('+coincide+'=1,'+total_tabla+'<>"NS",'+total_tabla+'<>"NA"),1,0),IF('+texto+'<>SUM('+NS+','+NA+'),1,0)'
            formulas = formula_aritmetica(
                total_tabla,
                Total,
                NS,
                NA,
                blancos,
                aha,
                coincide,
                hay,
                texto
                )
        
        alt = escribir(freal+2,formulas,hoja,letras)
        escr = ['=IF(SUM('+alt[1]+':'+alt[-3]+')>0,1,0)']
        donde=[alt[-1]]
        escribirgeneral(donde, escr, hoja)
        
        # print('alttttttttttttt', alt)
    
        # condi1 =[]
        # for l in letcol:
        #     condi1.append(lw(l,freal+(tupla[0]-fila),2))
    
        # for colu in condi1[1:]:
            
        #     hoja.conditional_formatting.add(colu,
        #                                     FormulaRule(formula=[condi1[0]+'="NA"'], stopIfTrue=True, fill=gris))
        try:
            fcon = condicional(letcol, freal, hoja, fila, tuplas)
            
            for i in fcon:
                hoja.conditional_formatting.add(i,
                                                FormulaRule(formula=[alt[-1]+'=1'], stopIfTrue=True, fill=redFill))
            # for i in fcon[1:]:
            #     hoja.conditional_formatting.add(i,
            #                                     FormulaRule(formula=[fcon[0]+'="NA"'], stopIfTrue=True, fill=gris))
        except:
            pass
        alt1.append(alt)
        fcon1.append(fcon)
    print('autosuma antes de funciones ', autosuma, iterar)
    if autosuma > 0 and iterar != 164:
        c = 0
        for letra in letcol:
            # autosumanormal([letra], freal, autosuma, hoja)
            autosumaportupla([letra], dfreal, autosuma, hoja, tuplas[c],fila)
            c += 1
    if men == 'si':  
        if autosuma > 0 and iterar != 164:
            menerror(freal+2,autosuma+2,hoja)
        if autosuma == 0 and iterar != 164:
            menerror(freal+2,freal+2,hoja)
    if autosuma == 0:
        
        if type(chek[0]) == list:
            nchek = []
            for c in chek:
                nchek += c
            chek = nchek
        dn = ','.join(chek)
        
        frame.agregar_valor_acolumna(dn,'coordenada')
    
    return [fcon1[0],alt1[0]]

def validarTS(fila,tuplas,freal,autosuma,hoja,letras):
    "Función parecida a variables pero para validar elementos que no están juntos en la tabla"
    print('argumentos de entrada funcion validarts',fila,tuplas,freal,autosuma )
    abc1 = list(string.ascii_uppercase)+['AA','AB','AC','AD']
    columnas = []
    letcol= []
    tuplas.sort()
    for tupla in tuplas:
        a = tupla[1]
        columnas.append(a)
    for i in columnas:
        s = abc1[i]
        letcol.append(s)
    alt1 = []
    fcon1 = []
    iterar = autosuma - freal - 1
    if iterar > 0:
        iterar += 1
    else:
        iterar = 1
    final_tabla.extraer([iterar])
    dfreal = freal + 0
    for i in range(0,iterar):
        freal = dfreal + i    
        
    
        coordss = crearcoordenada1(letcol, tuplas,fila,freal)
    
        val_des = coordss[1]
        total_tabla = coordss[0]
        
        fila_res = coordss[0]
        
        blancos = parablancos(coordss)
        aha = str(len(coordss))
        NS = 'COUNTIF('+fila_res+',"NS")'
        texto = 'COUNTIF('+fila_res+',"=*")'
        NA = 'COUNTIF('+fila_res+',"NA")'
        for i in coordss[1:]:
       
            NS += '+COUNTIF('+i+',"NS")'
            texto += '+COUNTIF('+i+',"=*")'
            NA += '+COUNTIF('+i+',"NA")'
        for i in coordss[2:]:
            val_des += ','+i
            
        Total = 'SUM('+val_des+')'
        coincide = 'IF(OR('+Total+'='+total_tabla+','+total_tabla+'="",AND('+total_tabla+'>0,'+NS+'='+str(len(coordss)-1)+')),0,1)'
        hay = 'IF(AND('+total_tabla+'=0,'+Total+'>0,'+NS+'=0),1,0),IF(OR(AND('+total_tabla+'="NS",'+Total+'>0),AND('+total_tabla+'="NS",'+NS+'<2)),1,0),=IF(AND('+total_tabla+'="NA",OR('+Total+'>0,'+NS+'>0)),1,0),IF(AND('+blancos+'>0,'+blancos+'<'+aha+','+total_tabla+'<>"NA"),1,0),IF(AND('+coincide+'=1,'+total_tabla+'<>"NS",'+total_tabla+'<>"NA"),1,0),IF('+texto+'<>SUM('+NS+','+NA+'),1,0)'
        formulas = formula_aritmetica(
            total_tabla,
            Total,
            NS,
            NA,
            blancos,
            aha,
            coincide,
            hay,
            texto
            )
        alt = escribir(freal+2,formulas,hoja,letras)
        escr = ['=IF(SUM('+alt[1]+':'+alt[-3]+')>0,1,0)']
        donde=[alt[-1]]
        escribirgeneral(donde, escr, hoja)
        try:
            fcon = condicional(letcol, freal, hoja, fila, tuplas)
            for i in fcon:
                hoja.conditional_formatting.add(i,
                                                FormulaRule(formula=[alt[-1]+'=1'], stopIfTrue=True, fill=redFill))
            # for i in fcon[1:]:
            #     hoja.conditional_formatting.add(i,
            #                                     FormulaRule(formula=[fcon[0]+'="NA"'], stopIfTrue=True, fill=gris))
        except:
            pass
        alt1.append(alt)
        fcon1.append(fcon)
        
    if autosuma > 0 and iterar != 165:
        c = 0
        for letra in letcol:
            autosumaportupla([letra], dfreal, autosuma, hoja, tuplas[c],fila)
            # autosumanormal([letra], freal, autosuma, hoja)
            c += 1
    if autosuma == 0:
        dn = ','.join(fcon1[0])
        frame.agregar_valor_acolumna(dn,'coordenada')
        
    # condi1 =[]
    # for l in letcol:
    #     condi1.append(lw(l,freal+(tupla[0]-fila),2))

    # for colu in condi1[1:]:
        
    #     hoja.conditional_formatting.add(colu,
    #                                     FormulaRule(formula=[condi1[0]+'="NA"'], stopIfTrue=True, fill=gris))

    
    return fcon1[0]


def validarTSD(fila,tuplas,freal,autosuma,hoja,letras,pregunta,codelitos):
    "Función para validar tablas de tipo de delito pero que no tienen elementos juntos en la tablaa"
    print('argumentos de entrada funcion validartsD',fila,tuplas,freal,autosuma )
    abc1 = list(string.ascii_uppercase)+['AA','AB','AC','AD']
    codelitos1=listalistas(codelitos)
    ntuplas = [] #hacer ajuste de las tuplas de otras filas
    copia_tuplas = [] 
    for tupla in tuplas:
        if tupla[0] != fila:
            a = (tupla[0]-1,tupla[1])
            copia_tuplas.append(a)
        if tupla[0] == fila:
            copia_tuplas.append(tupla)
        a = (tupla[0]-1,tupla[1])
        ntuplas.append(a)
    tuplas = ntuplas 
    columnas = []
    letcol= []
    tuplas.sort()
    for tupla in tuplas:
        a = tupla[1]
        columnas.append(a)
    for i in columnas:
        s = abc1[i]
        letcol.append(s)
        
    alt1 = []
    fcon1 = []
    iterar = autosuma - freal - 1
    if iterar > 0:
        iterar += 1
    else:
        iterar = 1
    
    dfreal = freal + 0
    for i in range(0,iterar):
        freal = dfreal + i 
    
        coordss = crearcoordenada1(letcol, tuplas,fila,freal)
    
        val_des = coordss[1]
        total_tabla = coordss[0]
        
        fila_res = coordss[0]
        
        blancos = parablancos(coordss)
        aha = str(len(coordss))
        NS = 'COUNTIF('+fila_res+',"NS")'
        texto = 'COUNTIF('+fila_res+',"=*")'
        NA = 'COUNTIF('+fila_res+',"NA")'
        for i in coordss[1:]:
       
            NS += '+COUNTIF('+i+',"NS")'
            texto += '+COUNTIF('+i+',"=*")'
            NA += '+COUNTIF('+i+',"NA")'
        for i in coordss[2:]:
            val_des += ','+i
            
        Total = 'SUM('+val_des+')'
        coincide = 'IF(OR('+Total+'='+total_tabla+','+total_tabla+'="",AND('+total_tabla+'>0,'+NS+'='+str(len(coordss)-1)+')),0,1)'
        hay = 'IF(AND('+total_tabla+'=0,'+Total+'>0,'+NS+'=0),1,0),IF(OR(AND('+total_tabla+'="NS",'+Total+'>0),AND('+total_tabla+'="NS",'+NS+'<2)),1,0),=IF(AND('+total_tabla+'="NA",OR('+Total+'>0,'+NS+'>0)),1,0),IF(AND('+blancos+'>0,'+blancos+'<'+aha+','+total_tabla+'<>"NA"),1,0),IF(AND('+coincide+'=1,'+total_tabla+'<>"NS",'+total_tabla+'<>"NA"),1,0),IF('+texto+'<>SUM('+NS+','+NA+'),1,0)'
        formulas = formula_aritmetica(
            total_tabla,
            Total,
            NS,
            NA,
            blancos,
            aha,
            coincide,
            hay,
            texto
            )
        alt = escribir(freal+2,formulas,hoja,letras)
        escr = ['=IF(SUM('+alt[1]+':'+alt[-3]+')>0,1,0)']
        donde=[alt[-1]]
        escribirgeneral(donde, escr, hoja)
        try:
            fcon = condicional(letcol, freal, hoja, fila, copia_tuplas)
            for i in fcon:
                hoja.conditional_formatting.add(i,
                                                FormulaRule(formula=[alt[-1]+'=1'], stopIfTrue=True, fill=redFill))
        except:
            pass
        alt1.append(alt)
        fcon1.append(fcon)
    if autosuma > 0:
        c = 0
        for letra in letcol:
            autosumadelito([letra],pregunta, hoja, tuplas[c])
            # autosumanormal([letra], freal, autosuma, hoja)
            c += 1
        
    
    
    
    formulas = ['coincide_con_desagregados?',
                'hay NA',
                'mayor al 25%']
    
       
    totls = totls1(codelitos1)
    delito_total=deltot(codelitos1,letcol[0],pregunta,tuplas[0],fila-1)
    sus_desagregados=desagregadoss(codelitos1,letcol[0],pregunta,tuplas[0],fila-1)
    az = []
    u = alt[-1]
    lc = u[:2]
    letrasesc= nl()
    we= letrasesc.index(lc)
    az.append(letrasesc[we+1])
    az.append(letrasesc[we+2])
    az.append(letrasesc[we+3])
    todo= []
    condi = [] #columnas ara condicionales, son 3 y ya va con fila real (Son las de la validación de la derecha de las tablas)
    for i in az:
        tr = i+str(dfreal+1)
        todo.append(tr)
        tr1 = i+str(dfreal+2)
        condi.append(tr1)
    escribirgeneral(todo,formulas, hoja)
    condi1 = [] #lista de letras que tienen los numeros de la fila
    condi2 = []
    c = 0
    for l in letcol:
        condi1.append(lw(l,dfreal+(tuplas[c][0]-fila-1)+2,2))
        condi2.append(lw(l,dfreal+(tuplas[c][0]-fila-1)+166,2)) #166 porque es ajuste de fin de tabla de delitos (164 + los dos que se suman de inicio)
        c += 1
    x_valida.extraer(condi)
    condi3 = []
    c = 0
    for letra in condi1:
        condi3.append(letra+':'+condi2[c])
        c += 1
    condi1 = condi3
    
    hoja.conditional_formatting.add(condi1[0],
                                    FormulaRule(formula=[condi[0]+'=1'], stopIfTrue=True, fill=redFill))
    hoja.conditional_formatting.add(condi1[0],
                                    FormulaRule(formula=[condi[2]+'=1'], stopIfTrue=True, fill=azul))
    for colu in condi1:
        
        hoja.conditional_formatting.add(colu,
                                        FormulaRule(formula=[condi[1]+'=1'], stopIfTrue=True, fill=gris))
        # if colu != condi1[0]:
        #     hoja.conditional_formatting.add(colu,
        #                                     FormulaRule(formula=[condi1[0]+'="NA"'], stopIfTrue=True, fill=gris))

    coin=az[0]
    ro=0
    for i in delito_total:
        totl=i
        low=sus_desagregados[ro]
        desag = low[0]+':'+low[-1]
        desa='SUM('+desag+')'
        #=SI(O(AS377=AT377,Y(AS377="NA",AT377=0),Y(AS377="NS",AT377=0)),SI(Y(AS377="NA",AT377=0),"NA",0),1)
        escri= ['=IF(OR('+totl+'=SUM('+desag+'),AND('+totl+'="NA",'+desa+'=0),AND('+totl+'="NS",'+desa+'=0),AND('+totl+'>0,SUM('+desag+')=0,COUNTIF('+desag+',"NS")>0)),IF(AND('+totl+'="NA",'+desa+'=0),"NA",0),1)']
        qw=totls[ro]+pregunta #es fila de pregunta no real
        ax=[coin+str(qw)]
        escribirgeneral(ax, escri, hoja)
        hoja.conditional_formatting.add(desag,
                                        FormulaRule(formula=['$'+coin+'$'+str(qw)+'=1'], stopIfTrue=True, fill=redFill))
        ro+=1
    #para segunda columna, de NA
    beta=az[1]
    listasdesagregados=desagregadoss(codelitos1, beta, pregunta,tuplas[0],tuplas[0][0]) #doble tupla para que d´cero, igual en la de abajjo ya que es para escribir en las columnas correspondientes
    gama=0
    for lista in listasdesagregados:
        b = delito_total[gama]
        for fila1 in lista:
            donde=[fila1]
            escr = ['=IF('+b+'="NA",1,0)']
            escribirgeneral(donde, escr, hoja)
        gama+=1
    
    delta=az[2]
    listasdesagregados1=desagregadoss(codelitos1, delta, pregunta, tuplas[0],tuplas[0][0])
    ya=0
    for lista in  listasdesagregados1:
        b = delito_total[ya] 
        ar = sus_desagregados[ya]
        ar1 = ar[-2]
        pri = [lista[-2]]
        escr = ['=IF('+ar1+'/'+b+'>0.25,1,0)']
        escribirgeneral(pri, escr, hoja)
        ya+=1
    escr = ['=IF(SUM('+alt[1]+':'+alt[-3]+')>0,1,0)'] #estas lineas son para formula de hay
    donde=[alt[-1]]
    escribirgeneral(donde, escr,hoja)   
    
    
    return fcon1[0]

def menerror(fila_inicio,fila_suma,hoja):
    "Función para poner mensaje de error. los argumentos ya deben tener su ajuste para la escritura"
    w = ['AF','AG','AH','AI','AJ','AK']
    r = ['1','2','4','4','5','6']
    c = 0
    for val in w:
        a = val+str(fila_inicio)
        b = val+str(fila_suma)
        if fila_inicio != fila_suma:
            que = ['=IF(SUM('+a+':'+b+')>0,'+r[c]+',0)']
            donde = [val+str(fila_suma+1)]
            escribirgeneral(donde, que,hoja)
        else:
            que = ['=IF('+a+'>0,'+r[c]+',0)']
            donde = [val+str(fila_suma+1)]
            escribirgeneral(donde, que,hoja)
        c+=1
    que1=['=SUM(AF'+str(fila_suma+1)+':'+'AH'+str(fila_suma+1)+')',
          '=SUM(AI'+str(fila_suma+1)+':'+'AK'+str(fila_suma+1)+')']
    donde1 = [ 'AH'+str(fila_suma+2), 'AK'+str(fila_suma+2)]
    escribirgeneral(donde1, que1,hoja)
    que2=['=CAMBIAR('+'AH'+str(fila_suma+2)+',0,"",1,"El total no puede ser cero si se registró un NS o un valor mayor que cero en los desagregados",2,"El total no puede ser NS si registraste algún valor o todos los desagregados son cero",3,"Error en total donde se indica cero y también donde se indica con NS",4,"Cuando el total es NA, la fila se debe dejar en blanco o llenarla con NA según sea la instrucción",5,"Hay error donde el total se indica como cero y también en total con NA",6,"Error donde el total es NS y también donde el total es NA",7,"En los totales que se indica cero, NS y NA ")',
          '=CAMBIAR('+'Ak'+str(fila_suma+2)+',0,"",4,"Hay espacios en blanco que deben ser llenados ",5,"La suma del total no coincide con los desagregados ",6,"Se registraron valores que no son aceptados como respuesta para esta pregunta",9,"Hay espacios en blanco y la suma no coincide con los desagregados",10,"Hay espacio en blanco y valores no admitidos",11,"La suma no coincide con los desagregados y se registran valores no aceptados",15,"Espacios en blanco, suma incorrecta y valores no aceptados")']
    donde2 = ['AH'+str(fila_suma+3), 'AK'+str(fila_suma+3)]  
    escribirgeneral(donde2, que2,hoja)
    
    return



def menerror_no_tab(fila_inicio,fila_suma,hoja,fin):
    "Poner mensaje de error en donde no es tabla. los argumentos ya deben tener su ajuste para la escritura"
    w = ['AF','AG','AH','AI','AJ','AK']
    r = ['1','2','4','4','5','6']
    c = 0
    for val in w:
        a = val+str(fila_inicio)
        b = val+str(fila_suma)
        if fila_inicio != fila_suma:
            que = ['=IF(SUM('+a+':'+b+')>0,'+r[c]+',0)']
            donde = [val+str(fila_suma+1)]
            escribirgeneral(donde, que,hoja)
        else:
            que = ['=IF('+a+'>0,'+r[c]+',0)']
            donde = [val+str(fila_suma+1)]
            escribirgeneral(donde, que,hoja)
        c+=1
    que1=['=SUM(AF'+str(fila_suma+1)+':'+'AH'+str(fila_suma+1)+')',
          '=SUM(AI'+str(fila_suma+1)+':'+'AK'+str(fila_suma+1)+')']
    donde1 = [ 'AH'+str(fila_suma+2), 'AK'+str(fila_suma+2)]
    escribirgeneral(donde1, que1,hoja)
    que2=['=CAMBIAR('+'AH'+str(fila_suma+2)+',0,"",1,"El total no puede ser cero si se registró un NS o un valor mayor que cero en los desagregados",2,"El total no puede ser NS si registraste algún valor o todos los desagregados son cero",3,"Error en total donde se indica cero y también donde se indica con NS",4,"Cuando el total es NA, los desagregados deben quedar en blanco o llenarla con NA según sea la instrucción",5,"Hay error donde el total se indica como cero y también en total con NA",6,"Error donde el total es NS y también donde el total es NA",7,"En los totales que se indica cero, NS y NA ")',
          '=CAMBIAR('+'Ak'+str(fila_suma+2)+',0,"",4,"Hay espacios en blanco que deben ser llenados ",5,"La suma del total no coincide con los desagregados ",6,"Se registraron valores que no son aceptados como respuesta para esta pregunta",9,"Hay espacios en blanco y la suma no coincide con los desagregados",10,"Hay espacio en blanco y valores no admitidos",11,"La suma no coincide con los desagregados y se registran valores no aceptados",15,"Espacios en blanco, suma incorrecta y valores no aceptados")']
    donde2 = ['AH'+str(fila_suma+3), 'AK'+str(fila_suma+3)]  
    escribirgeneral(donde2, que2,hoja)
    try: #escribir debajo de la tabla los mensajes de error
       que3 = ['='+donde2[0],'='+donde2[1]] 
       donde3 = ['C'+str(fin+1),'C'+str(fin+2)]
       escribir_mensajes_error(donde3, que3,hoja)
    except:
        pass
    return




def parablancos(coordss):
    "Crear formula de blancos"
    formula = 'COUNTBLANK('+coordss[0]+')'
    ncor = coordss[1:]
    c=0
    for i in ncor:
        a = '+COUNTBLANK('+ncor[c]+')'
        formula = formula+a
        c+=1
    return formula

def paravalerror(coordss):
    "Crear formula para valor error"
    formula = 'COUNTIF('+coordss[0]+',"=*")-COUNTIF('+coordss[0]+',"NA")-COUNTIF('+coordss[0]+',"NS")'
    ncor = coordss[1:]
    c=0
    for i in ncor:
        a = '+COUNTIF('+ncor[c]+',"=*")-COUNTIF('+ncor[c]+',"NA")-COUNTIF('+ncor[c]+',"NS")'
        formula = formula+a
        c+=1
    return formula

def crearcoordenada(listaletras,freal):
    "Crear coordenada tipo excel"
    # print('este es crear coordenada',listaletras,freal)
    res = []
    for i in listaletras:
        a = i+str(freal+2)
        res.append(a)
    return res

def crearcoordenada1(listaletras,tuplas,fila, freal):
    "Crear coordenada tipo excel, con ajustes"
    # print('este es crear coordenada',listaletras,freal)
    r = freal-fila
    fs = []
    for tup in tuplas:
        a = tup[0] + r
        fs.append(a)
        
    res = []
    c = 0
    for i in listaletras:
        n = fs[c]+2
        if n == freal+2:
            a = i+str(n)
        else:
            a = i+str(n+1)
        res.append(a)
        c += 1
    return res

abc = list(string.ascii_uppercase)


def hay1(lista,actual):
    "Crear rango de coordenadas tipo excel"
    a = lista.index(actual)
    b = lista[a-6]
    c = lista[a-1]
    d = b+':'+c
    return d




def letras(inicio,extencion):
    "generar lista con combinacion de letras para celdas inicio es AF normalmente"
    tr = list(inicio)
    ind = []
    for i in tr:
        v = abc.index(i)
        ind.append(v)
    v2 = ind[1]
    lista = []
    for i in range(extencion):
        d = v2+i
        c = abc[d]
        a =tr[0]+c
        lista.append(a)
    
    return lista

def nl ():
    "Gnerar lista de lestras para coordenas en donde se van a imprimir las validaciones"
    rango = ['A','B','C','D','E','F','G','H','I','J','K','L']
    res = []
    for letra in rango:
        for lt in abc:
            e = letra+lt
            res.append(e)
    return res[5:]

def letraslide(numero):
    "Generar lista de letras a tamaño necesario"
    ldeletras = nl()
    indletras = [0,10,20,30,40,50,60,70,80,90,100,110,120,130,140,150,160,
                 170,180,190,200,210,220,230,240,250]
    ind = indletras[numero+1]
    ind0 =indletras[numero]
    res = ldeletras[ind0:ind]
    return res


def escribir(fila,formulas,hoja,letras):
    "Función para escribir en el archivo de excel"
    y=[]
    con = 0
    for i in formulas:
        filas = [fila-1,fila]
        datos = [i,formulas[i]]
        c = 0        
        for fila in filas:
            letr = letras[con]+str(fila)
            if fila == filas[0]:
                y.append(letr)
               
            else:  
                hoja[letr] =datos[c] 
                y.append(letr)
            c+=1
        con+=1
    return y



def escribirgeneral(donde,que,hoja):
    "dodne son las letras con el numero de fila y que es lo que se va a escribir, ambos listas"
    l = 0 
    for i in donde:
        hoja[i] = que[l]
        l+=1
    return

def escribir_mensajes_error(donde,que,hoja):
    "dodne son las letras con el numero de fila y que es lo que se va a escribir, ambos listas"
    l = 0 
    tipo = Font(name = 'Arial', size = 10, color = 'EE1111', bold = True )
    alin = Alignment(wrap_text=False)
    for i in donde:
        a1 = hoja[i]
        a1.font = tipo
        a1.alignment = alin
        hoja[i] = que[l]
        l+=1
    return






